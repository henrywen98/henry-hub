"""
_common.py — 三个 extractor 共享的纯结构 + 共用样式工具。

设计原则: 只放真正"行为一致"的工具,不强行统一不同脚本的语义。

入选 (单一权威实现):
- iter_block_items / normalize / heading_level: 三脚本中 (或 agg+sol 中) 实现 byte-equivalent
- 样式常量 HEADER_FILL/HEADER_FONT/ZEBRA_FILL/THIN_BORDER + _font + _apply_header +
  _apply_data_styles: extract_solution.py + extract_quotation.py 视觉规范一致
  (深蓝表头 / 浅灰斑马 / 浅灰边框)。aggregate.py 有自己的多色样式系统 (手填列橙/
  警告/成功),不在这里;它只用上面的 3 个结构工具。
- 报价列浅黄 PRICE_FILL 仅 extract_quotation 使用,通过 _apply_data_styles 的
  price_col 参数控制 (默认 None 不高亮)。

不入选 (语义差异):
- find_col: aggregate 用完全匹配 `if h in candidates`,extract_solution 用子串匹配
  `if cand in h`,extract_quotation 把子串匹配 inline 在 map_columns 里。
- table_header: aggregate 用 `normalize(c.text)`,extract_solution 用 cell_text
  (多 paragraph join)。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Iterator

from docx.document import Document as DocxDocumentT
from docx.oxml.ns import qn
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# docx 结构工具
# ---------------------------------------------------------------------------


def iter_block_items(doc: DocxDocumentT) -> Iterator[Paragraph | DocxTable]:
    """按文档顺序产出段落或表格。"""
    body = doc.element.body
    for child in body.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, doc)
        elif child.tag == qn("w:tbl"):
            yield DocxTable(child, doc)


def normalize(text: str) -> str:
    """空白折叠 + strip。"""
    return re.sub(r"\s+", " ", (text or "").strip())


def heading_level(p: Paragraph) -> int | None:
    """从 paragraph style 解析 Heading 层级。

    匹配 "Heading 1" / "Heading10" / "heading 2" 等变体 (大小写不敏感)。
    """
    style = (p.style.name if p.style else "") or ""
    m = re.match(r"Heading\s*(\d+)", style, re.IGNORECASE)
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# 共享 openpyxl 样式 (extract_solution.py + extract_quotation.py)
# ---------------------------------------------------------------------------


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # 深蓝
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
PRICE_FILL = PatternFill("solid", fgColor="FFF2CC")  # 报价列浅黄 (extract_quotation 专用)
MANUAL_HEADER_FILL = PatternFill("solid", fgColor="FFC000")  # 手填评估列橙色表头
_SIDE = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


def _font(bold: bool = False, size: int = 10, color: str = "000000") -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)


def _apply_header(
    ws: Worksheet,
    headers: list[str],
    manual_col_indices: list[int] | None = None,
) -> None:
    """深蓝表头 + 居中 + 冻结首行 + 自动筛选。

    manual_col_indices: 1-based 列号列表,这些列的表头改用橙色高亮 (评估手填列)。
    """
    manual_set = set(manual_col_indices or [])
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = MANUAL_HEADER_FILL if i in manual_set else HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _apply_data_styles(
    ws: Worksheet,
    n_rows: int,
    n_cols: int,
    price_col: int | None = None,
) -> None:
    """斑马纹数据行 + 浅灰边框。可选高亮 price_col 列为浅黄。"""
    for r in range(2, n_rows + 2):
        zebra = ZEBRA_FILL if r % 2 == 0 else None
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.font = _font(size=10)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = THIN_BORDER
            if col == price_col:
                c.fill = PRICE_FILL
            elif zebra:
                c.fill = zebra
        ws.row_dimensions[r].height = 30


# ---------------------------------------------------------------------------
# 手填列保留 (Mode B/C 重跑不覆盖评估列)
# ---------------------------------------------------------------------------


def load_manual_columns(
    xlsx_path: Path,
    sheet_name: str,
    key_cols: tuple[str, ...],
    manual_cols: tuple[str, ...],
    header_row: int = 1,
) -> dict[tuple, dict[str, object]]:
    """从已有 xlsx 的指定 sheet 读取手填列,按主键索引返回。

    用于 Mode B/C 重跑时保留用户手工填写的评估列 (复用度/成熟度/已应用项目/
    维护责任人/备注 等)。aggregate.py (Mode A) 用自己的 load_existing_manual,
    那个有三层主键合并 + 多版本兼容,不走这里。

    Args:
        xlsx_path: 旧 xlsx 路径,不存在 / 读不开 / sheet 缺失 → 返回 {}
        sheet_name: 要读的 sheet 名
        key_cols: 主键列名 (顺序敏感),共同形成 dict key
        manual_cols: 要保留的手填列名;只保留有非空值的列
        header_row: header 所在行号 (默认 1)

    Returns:
        {(key1_str, key2_str, ...): {manual_col: value, ...}, ...}
    """
    if not xlsx_path.exists():
        return {}
    try:
        wb = load_workbook(str(xlsx_path), data_only=True)
    except Exception as e:
        print(f"[警告] 无法读取旧 Excel ({e}),跳过手填列保留", file=sys.stderr)
        return {}
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header = [c.value for c in ws[header_row]]
    idx = {h: i for i, h in enumerate(header) if h}

    missing_keys = [k for k in key_cols if k not in idx]
    if missing_keys:
        print(
            f"[警告] 旧 xlsx '{sheet_name}' sheet 缺主键列 {missing_keys},跳过手填合并",
            file=sys.stderr,
        )
        return {}

    out: dict[tuple, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        try:
            key = tuple(
                "" if row[idx[k]] is None else str(row[idx[k]]) for k in key_cols
            )
        except IndexError:
            continue
        manual_values: dict[str, object] = {}
        for col in manual_cols:
            i = idx.get(col)
            if i is None or i >= len(row):
                continue
            v = row[i]
            if v is not None and str(v).strip() != "":
                manual_values[col] = v
        if manual_values:
            out[key] = manual_values
    return out
