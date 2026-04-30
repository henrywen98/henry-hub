#!/usr/bin/env python3
"""
extract_quotation.py — 从「报价方案 / 功能清单」.xlsx 抽取功能模块,生成统一格式的 Excel。

跟 aggregate.py 的区别:
- aggregate.py 输入是 .docx PRD,本脚本输入是 .xlsx 报价表
- 报价表通常多 sheet (一个客户一个 sheet),每行 = 一个功能子模块
- 列结构典型: [序号, 业务范围, 模块, 子模块, 功能描述, 分摊报价] (+ 可选 优先级/备注)
- 也兼容无价的"招标功能"sheet: [序号, 业务范围, 功能模块, 二级功能, 功能描述]

输出 2 sheet:
- 模块报价: 跨所有源 sheet 合并的功能行,统一字段 [来源sheet, 业务范围, 模块, 子模块, 功能描述, 报价, 优先级, 备注]
- README

使用:
    python extract_quotation.py --input <quotation.xlsx> --output <out>.xlsx
    python extract_quotation.py --input <quotation.xlsx> --sheet "杭州产投-报价清单"   # 只抽指定 sheet
    python extract_quotation.py --input <quotation.xlsx> --inspect                       # 列出 sheet 与表头
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from _common import _apply_data_styles, _apply_header, _font, load_manual_columns


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------


@dataclass
class QuotationRow:
    source_sheet: str
    business_scope: str
    module: str
    sub_module: str
    description: str
    price: str
    priority: str
    # 评估手填列 — 重跑时通过 load_manual_columns 保留
    reuse_level: str = ""        # 复用度
    maturity: str = ""           # 成熟度
    applied_projects: str = ""   # 已应用项目
    maintainer: str = ""         # 维护责任人
    note: str = ""               # 备注


# ---------------------------------------------------------------------------
# 表头识别
# ---------------------------------------------------------------------------

# 不同 sheet 用的列名近义词
COL_PATTERNS = {
    "scope": ("业务范围", "系统", "业务领域", "范围"),
    "module": ("功能模块", "模块"),
    "sub": ("子模块", "二级功能", "功能", "功能名称", "子功能"),
    "desc": ("功能描述", "功能事项说明", "描述", "说明", "事项说明"),
    "price": ("分摊报价", "报价", "金额", "费用", "单价", "价格", "小计"),
    "prio": ("优先级",),
    "note": ("备注", "remark"),
}

SKIP_ROW_TOKENS = ("合计", "小计", "总计", "Total")


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def find_header_row(ws: Worksheet, max_scan: int = 10) -> tuple[int, list[str]] | None:
    """扫前 max_scan 行,找出"看起来像表头"的那一行。

    判定: 至少命中 1 个"模块/功能描述/报价"关键字,且至少 4 个非空单元格。
    """
    for r in range(1, min(max_scan, ws.max_row) + 1):
        cells = [_norm(c.value) for c in ws[r]]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 4:
            continue
        joined = " ".join(non_empty)
        # 必须命中关键字之一
        if any(p in joined for pats in (COL_PATTERNS["module"], COL_PATTERNS["desc"]) for p in pats):
            return r, cells
    return None


def map_columns(header: list[str]) -> dict[str, int]:
    """把表头映射到 {字段名: 列索引}。每个字段取第一个匹配的列。"""
    m: dict[str, int] = {}
    for field_name, patterns in COL_PATTERNS.items():
        for i, h in enumerate(header):
            if not h:
                continue
            if any(p in h for p in patterns):
                if field_name not in m:
                    m[field_name] = i
                    break
    return m


# ---------------------------------------------------------------------------
# 抽取主逻辑
# ---------------------------------------------------------------------------


def extract_sheet(ws: Worksheet) -> list[QuotationRow]:
    """从一个 sheet 抽出功能行。"""
    located = find_header_row(ws)
    if located is None:
        return []
    header_row, header = located
    cols = map_columns(header)
    if "module" not in cols and "desc" not in cols and "sub" not in cols:
        # 找不到核心列,跳过
        return []

    # 处理 merged cells 的"父值继承"(比如"业务范围"列只在第一行写一次,后面留空)
    last_values: dict[str, str] = {}
    rows: list[QuotationRow] = []

    for r_idx in range(header_row + 1, ws.max_row + 1):
        cells = [_norm(c.value) for c in ws[r_idx]]
        if not any(cells):
            continue

        # 提取每个字段
        def get(field_name: str) -> str:
            idx = cols.get(field_name)
            if idx is None or idx >= len(cells):
                return ""
            return cells[idx]

        row_data = {f: get(f) for f in COL_PATTERNS}

        # 跳过合计行: 模块/子模块/描述任一含"合计/小计"
        joined = " ".join(row_data.values())
        if any(tok in joined for tok in SKIP_ROW_TOKENS):
            continue

        # 父值继承(scope / module 列经常是 merged cell,后面行留空)
        for fk in ("scope", "module"):
            if row_data[fk]:
                last_values[fk] = row_data[fk]
            else:
                row_data[fk] = last_values.get(fk, "")

        # 至少要有 desc 或 sub 的内容才算一行
        if not (row_data["desc"] or row_data["sub"]):
            continue

        rows.append(QuotationRow(
            source_sheet=ws.title,
            business_scope=row_data["scope"],
            module=row_data["module"],
            sub_module=row_data["sub"],
            description=row_data["desc"],
            price=row_data["price"],
            priority=row_data["prio"],
            note=row_data["note"],
        ))

    return rows


def extract_from_xlsx(path: Path, sheet_filter: str | None) -> list[QuotationRow]:
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        print(f"无法打开 xlsx ({path}): {e}", file=sys.stderr)
        return []
    out: list[QuotationRow] = []
    for ws in wb.worksheets:
        if sheet_filter and sheet_filter not in ws.title:
            continue
        rows = extract_sheet(ws)
        out.extend(rows)
    return out


# ---------------------------------------------------------------------------
# Excel 输出
# ---------------------------------------------------------------------------


def write_quotation_sheet(ws: Worksheet, rows: list[QuotationRow]):
    headers = [
        "序号", "来源 sheet", "业务范围", "模块", "子模块", "功能描述", "报价", "优先级",
        # 评估手填列 (橙色表头)
        "复用度", "成熟度", "已应用项目", "维护责任人", "备注",
    ]
    manual_indices = [9, 10, 11, 12, 13]
    _apply_header(ws, headers, manual_col_indices=manual_indices)
    for i, q in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=q.source_sheet)
        ws.cell(row=i, column=3, value=q.business_scope)
        ws.cell(row=i, column=4, value=q.module)
        ws.cell(row=i, column=5, value=q.sub_module)
        ws.cell(row=i, column=6, value=q.description)
        ws.cell(row=i, column=7, value=q.price)
        ws.cell(row=i, column=8, value=q.priority)
        ws.cell(row=i, column=9, value=q.reuse_level)
        ws.cell(row=i, column=10, value=q.maturity)
        ws.cell(row=i, column=11, value=q.applied_projects)
        ws.cell(row=i, column=12, value=q.maintainer)
        ws.cell(row=i, column=13, value=q.note)
    _apply_data_styles(ws, len(rows), len(headers), price_col=7)
    widths = [6, 22, 16, 22, 22, 50, 14, 10, 12, 12, 22, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_readme_sheet(ws: Worksheet, source: Path, sheet_filter: str | None,
                       n_rows: int, source_sheets: list[str]):
    lines = [
        ("功能报价清单 · README", True),
        ("", False),
        ("用途", True),
        ("从 .xlsx 报价表 / 功能清单 抽取功能模块,合并到一张统一表里。", False),
        ("", False),
        ("Sheet 说明", True),
        ("• 模块报价 — 每行 = 一个功能子模块,带 报价/优先级 列。", False),
        ("    含 5 个橙色手填评估列: 复用度 / 成熟度 / 已应用项目 / 维护责任人 / 备注。", False),
        ("    重跑时会按 (来源 sheet, 业务范围, 模块, 子模块) 主键合并保留手填值,不会被覆盖。", False),
        ("• README — 本说明。", False),
        ("", False),
        ("本次抽取参数", True),
        (f"• 源文件: {source.name}", False),
        (f"• Sheet 筛选: {sheet_filter or '(全部 sheet)'}", False),
        (f"• 抽出行数: {n_rows}", False),
        (f"• 涵盖 sheet: {', '.join(source_sheets) if source_sheets else '(无)'}", False),
        ("", False),
        ("再跑命令", True),
        (f"python extract_quotation.py --input {source.name} --output <out>.xlsx", False),
        ("", False),
        ("提示", True),
        ("• 如果一个客户一份 sheet,可用 --sheet '客户名' 只抽特定 sheet", False),
        ("• 报价列(浅黄高亮)是金钱信息,在做跨项目复用判断时是关键", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = _font(bold=bold, size=12 if bold else 10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 22 if bold else 18
    ws.column_dimensions["A"].width = 100


_MANUAL_COLS = ("复用度", "成熟度", "已应用项目", "维护责任人", "备注")
_MANUAL_FIELD_MAP = {
    "复用度": "reuse_level",
    "成熟度": "maturity",
    "已应用项目": "applied_projects",
    "维护责任人": "maintainer",
    "备注": "note",
}


def _merge_manual_into_rows(rows: list[QuotationRow], output: Path) -> None:
    """从已有 xlsx 读取手填列,按 (来源 sheet, 业务范围, 模块, 子模块) 合并回新 rows。"""
    manual = load_manual_columns(
        output,
        sheet_name="模块报价",
        key_cols=("来源 sheet", "业务范围", "模块", "子模块"),
        manual_cols=_MANUAL_COLS,
    )
    if not manual:
        return
    new_keys = {(r.source_sheet, r.business_scope, r.module, r.sub_module) for r in rows}
    matched = 0
    for r in rows:
        key = (r.source_sheet, r.business_scope, r.module, r.sub_module)
        vals = manual.get(key)
        if not vals:
            continue
        matched += 1
        # 手填值优先: Mode C 的 "备注" 列是双重身份 (源 xlsx 抽出 + 用户评估手填)。
        # 用户一旦动过这列,重跑必须保留用户值。空字符串视为"无填写",不会进 vals
        # (load_manual_columns 已过滤),所以源备注在用户没碰过时仍可见。
        for col, attr in _MANUAL_FIELD_MAP.items():
            if col in vals:
                setattr(r, attr, str(vals[col]))
    orphans = [k for k in manual if k not in new_keys]
    print(
        f"[手填保留] 合并 {matched}/{len(manual)} 行手填值"
        + (f";丢弃 {len(orphans)} 个孤儿: {orphans[:3]}" if orphans else ""),
        file=sys.stderr,
    )


def write_excel(rows: list[QuotationRow], source: Path, sheet_filter: str | None, output: Path):
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    write_quotation_sheet(wb.create_sheet("模块报价"), rows)
    source_sheets = sorted({r.source_sheet for r in rows})
    write_readme_sheet(wb.create_sheet("README"), source, sheet_filter, len(rows), source_sheets)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description="报价方案 / 功能清单 .xlsx 抽取")
    ap.add_argument("--input", required=True, type=Path, help="输入 .xlsx 路径")
    ap.add_argument("--output", type=Path, default=None, help="输出 .xlsx (默认同目录同名)")
    ap.add_argument("--sheet", type=str, default=None, help="只抽 sheet 名包含此串的 sheet")
    ap.add_argument("--inspect", action="store_true", help="只列出 sheet 与表头")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"找不到输入: {args.input}", file=sys.stderr)
        return 2

    if args.inspect:
        try:
            wb = load_workbook(str(args.input), read_only=True, data_only=True)
        except Exception as e:
            print(f"无法打开 xlsx ({args.input}): {e}", file=sys.stderr)
            return 2
        print(f"Sheets ({len(wb.sheetnames)}):")
        for ws in wb.worksheets:
            located = find_header_row(ws)
            if located:
                hr, header = located
                cols = map_columns(header)
                fields = [k for k in COL_PATTERNS if k in cols]
                print(f"  • {ws.title}: header@R{hr}, 命中字段={fields}")
            else:
                print(f"  • {ws.title}: 未找到表头")
        return 0

    rows = extract_from_xlsx(args.input, args.sheet)
    print(f"抽出 {len(rows)} 行,涵盖 sheet: {sorted({r.source_sheet for r in rows})}")

    output = args.output if args.output else args.input.with_suffix(".功能模块.xlsx")
    _merge_manual_into_rows(rows, output)
    write_excel(rows, args.input, args.sheet, output)
    print(f"已生成: {output}")

    if not rows:
        print(
            "[警告] 所有 sheet 都未识别到表头/数据,可能 xlsx 结构非标准 "
            "(用 --inspect 列出 sheet 与表头命中)",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
