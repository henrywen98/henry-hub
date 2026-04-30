"""程序化对单个 run 的 outputs/ 做 assertions 评分。

用法:
  python grade_run.py <run_dir>
其中 run_dir 内含 outputs/ 子目录(产物)与 ../eval_metadata.json(读断言)。

输出: 在 run_dir 内写 grading.json (字段: expectations[{text,passed,evidence}])
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Assertion checkers
# ---------------------------------------------------------------------------

FEATURE_LIST_KEYS = ("FR", "功能", "需求", "feature", "requirement", "spec")
FIELD_DEF_KEYS = ("字段", "field", "key", "type", "属性")
LIFECYCLE_KEYS = ("生命周期", "状态", "lifecycle", "status", "PRD", "签字", "确认", "缺")
MANUAL_COL_KEYS = ("开发状态", "责任人", "优先级", "计划交付", "owner", "assignee", "priority", "status", "due")


def _all_xlsx(d: Path) -> list[Path]:
    return sorted([p for p in d.rglob("*.xlsx") if not p.name.startswith("~$")])


def _flatten_headers(wb) -> list[tuple[str, list[str]]]:
    """对每个 sheet 取前 3 行的所有非空文本作为表头候选。"""
    out = []
    for s in wb.sheetnames:
        ws = wb[s]
        candidates = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 3), values_only=True):
            for c in row:
                if c is None:
                    continue
                candidates.append(str(c))
        out.append((s, candidates))
    return out


def _has_key(headers: list[str], keys: tuple[str, ...]) -> tuple[bool, str]:
    for h in headers:
        for k in keys:
            if k.lower() in h.lower():
                return True, f"header={h!r} hit={k!r}"
    return False, ""


def check(run_dir: Path, assertions: list[dict]) -> list[dict]:
    out_dir = run_dir / "outputs"
    xlsx_files = _all_xlsx(out_dir) if out_dir.exists() else []

    # 加载第一个 xlsx(若有)
    wb = None
    headers_per_sheet: list[tuple[str, list[str]]] = []
    if xlsx_files:
        try:
            wb = load_workbook(xlsx_files[0], data_only=True)
            headers_per_sheet = _flatten_headers(wb)
        except Exception:
            wb = None
            headers_per_sheet = []

    results: list[dict] = []
    for a in assertions:
        aid = a.get("id", "")
        text = a.get("text", "")
        passed, evidence = _eval_assertion(aid, xlsx_files, wb, headers_per_sheet)
        results.append({"text": text, "passed": passed, "evidence": evidence})
    return results


def _eval_assertion(
    aid: str,
    xlsx_files: list[Path],
    wb,
    headers_per_sheet: list[tuple[str, list[str]]],
) -> tuple[bool, str]:
    if aid == "produces_xlsx":
        return (len(xlsx_files) > 0, f"found {len(xlsx_files)} xlsx: {[p.name for p in xlsx_files]}")

    if not xlsx_files or wb is None:
        return False, "no xlsx loadable"

    if aid == "multi_sheet":
        n = len(wb.sheetnames)
        return (n >= 3, f"{n} sheets: {wb.sheetnames}")

    if aid == "has_feature_list":
        for sname, headers in headers_per_sheet:
            ok, ev = _has_key(headers, FEATURE_LIST_KEYS)
            if ok:
                return True, f"sheet={sname} {ev}"
        return False, "no sheet matched feature-list keys"

    if aid == "has_capability_sheet":
        # 「能力清单」一般在 Sheet 1。允许出现在前两张里(兼容兼容性变体)。
        for sname in wb.sheetnames[:2]:
            if "能力" in sname or "capabilit" in sname.lower():
                ws = wb[sname]
                data_rows = max(0, ws.max_row - 2)
                if data_rows <= 15:
                    return True, f"sheet={sname} rows={data_rows}"
                return False, f"sheet={sname} 行数 {data_rows} 太多,能力点该粗颗粒"
        return False, "前两张 sheet 都没找到「能力清单」"

    if aid == "has_module_overview":
        # 「业务模块概览」级别(~16 行子模块,而不是 ~80 行 FR)。
        # 当前布局下一般是 Sheet 2(Sheet 1 是「能力清单」)。允许 Sheet 1 或 Sheet 2。
        for sname in wb.sheetnames[:2]:
            if "业务模块" in sname or "模块概览" in sname or "module" in sname.lower():
                ws = wb[sname]
                data_rows = max(0, ws.max_row - 2)
                if data_rows <= 30:
                    return True, f"sheet={sname} rows={data_rows}"
                return False, f"sheet={sname} 行数 {data_rows} 太多,可能仍是 FR 级"
        return False, f"前两张 sheet '{wb.sheetnames[:2]}' 都不像业务模块概览"

    if aid == "has_field_definitions":
        for sname, headers in headers_per_sheet:
            ok, ev = _has_key(headers, FIELD_DEF_KEYS)
            if ok:
                return True, f"sheet={sname} {ev}"
        return False, "no sheet matched field-def keys"

    if aid == "has_lifecycle_or_status":
        for sname, headers in headers_per_sheet:
            ok, ev = _has_key(headers, LIFECYCLE_KEYS)
            if ok:
                return True, f"sheet={sname} {ev}"
        return False, "no sheet matched lifecycle/status keys"

    if aid in ("coverage_01_03", "coverage_05", "coverage_08"):
        prefix = {"coverage_01_03": "01-03", "coverage_05": "05", "coverage_08": "08"}[aid]
        count = 0
        sample = ""
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and str(cell).strip().startswith(prefix):
                        count += 1
                        if not sample:
                            sample = f"sheet={sname} row sample"
                        break
        return (count > 0, f"{count} cells/rows match prefix {prefix}; {sample}")

    if aid == "manual_columns_present":
        for sname, headers in headers_per_sheet:
            ok, ev = _has_key(headers, MANUAL_COL_KEYS)
            if ok:
                return True, f"sheet={sname} {ev}"
        return False, "no sheet matched manual-column keys"

    if aid == "has_styling":
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, min(ws.max_row, 4) + 1):
                for c in range(1, min(ws.max_column, 20) + 1):
                    cell = ws.cell(row=r, column=c)
                    fill = cell.fill
                    if fill and fill.patternType and fill.fgColor:
                        rgb = fill.fgColor.rgb
                        if rgb and rgb not in ("00000000", "FFFFFFFF", None):
                            return True, f"sheet={sname} cell={cell.coordinate} fill={rgb}"
        return False, "no styled headers detected"

    if aid == "has_freeze_panes":
        for sname in wb.sheetnames:
            ws = wb[sname]
            if ws.freeze_panes and ws.freeze_panes != "A1":
                return True, f"sheet={sname} freeze_panes={ws.freeze_panes}"
        return False, "no freeze_panes set"

    if aid == "recalc_zero_errors":
        # 调 LibreOffice recalc.py;路径优先级: 环境变量 > 自动发现 > 旧硬编码
        import os
        recalc = os.environ.get("XLSX_RECALC_PY", "")
        if not recalc or not Path(recalc).exists():
            # 在 plugin cache 下找(版本 hash 会变)
            cache_root = Path.home() / ".claude" / "plugins" / "cache"
            candidates = list(cache_root.glob("*/skills/xlsx/scripts/recalc.py")) \
                       + list(cache_root.glob("**/document-skills/*/skills/xlsx/scripts/recalc.py"))
            if candidates:
                recalc = str(candidates[0])
        if not recalc or not Path(recalc).exists():
            return False, "recalc.py not found (set $XLSX_RECALC_PY or install document-skills:xlsx)"
        try:
            r = subprocess.run(
                ["python3", recalc, str(xlsx_files[0]), "60"],
                capture_output=True, text=True, timeout=120,
            )
            try:
                data = json.loads(r.stdout)
            except Exception:
                return False, f"non-JSON output: {r.stdout[:200]}"
            ok = data.get("status") == "success" and data.get("total_errors", 1) == 0
            return ok, f"recalc: {data}"
        except Exception as e:
            return False, f"recalc raised: {e}"

    return False, f"unknown assertion id: {aid}"


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: grade_run.py <run_dir>", file=sys.stderr)
        return 1
    run_dir = Path(sys.argv[1])
    metadata_path = None
    for candidate in (run_dir.parent, run_dir.parent.parent, run_dir.parent.parent.parent):
        p = candidate / "eval_metadata.json"
        if p.exists():
            metadata_path = p
            break
    if metadata_path is None:
        print(f"[{run_dir}] no eval_metadata.json found in any ancestor", file=sys.stderr)
        return 1
    meta = json.loads(metadata_path.read_text())
    assertions = meta.get("assertions", [])
    results = check(run_dir, assertions)
    out = {"expectations": results}
    grading_path = run_dir / "grading.json"
    grading_path.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    passed = sum(1 for r in results if r["passed"])
    print(f"[{run_dir.name}] {passed}/{len(results)} passed -> {grading_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
