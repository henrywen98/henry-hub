#!/usr/bin/env python3
"""
extract_solution.py — 从「解决方案 / 建设方案 / 升级方案」.docx 抽取功能模块,生成 Excel。

跟 aggregate.py 的区别:
- aggregate.py 假设输入是有 FR 编号表的 PRD,有"前缀方案 / 能力清单 / AI 字段定义"等强结构
- 本脚本面向解决方案文档: 章节嵌套深、没有 FR 编号、模块以 Heading N 标题为载体、
  内嵌的"功能模块 / 费用 / 工作量"表是关键金矿

输出 3 sheet:
- 模块概览: 一行 = 一个功能模块(标题 + 章节描述 + 内嵌表抽到的费用/工时,如有)
- 章节明细: 每段 paragraph 一行,保留原文档结构便于审阅
- README

使用:
    python extract_solution.py --input <doc.docx> --output <out.xlsx> [--heading-level 3]
    python extract_solution.py --input <doc.docx>                       # 自动选 heading level
    python extract_solution.py --inspect <doc.docx>                     # 只看 heading 分布
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from docx import Document as DocxDocument
from docx.document import Document as DocxDocumentT
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from _common import (
    _apply_data_styles,
    _apply_header,
    _font,
    heading_level,
    iter_block_items,
    normalize,
)


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------


@dataclass
class ModuleEntry:
    """功能模块 — 一行模块概览。"""

    index: int
    title: str
    parent_path: str  # 从顶层到本模块的章节路径(如 "升级内容 > 核心业务功能完善")
    description: str  # 模块下的 paragraph 文本汇总(裁剪)
    estimated_cost: str = ""    # 内嵌"费用"表抽出
    estimated_effort: str = ""  # 内嵌"工作量"表抽出
    note: str = ""


@dataclass
class SectionEntry:
    """章节明细 — 一行 paragraph。"""

    module_title: str    # 所属模块
    heading_level: int   # 0 = paragraph (非 heading)
    text: str


# ---------------------------------------------------------------------------
# docx 工具
# ---------------------------------------------------------------------------


def detect_heading_level(doc) -> int:
    """Auto-pick Heading level for "feature module" — 模块层级通常是 heading 数量分布的次峰。

    简化策略:
    - 统计每层 heading 数量
    - 模块层应该 >= 5 个 heading,且不在最深层
    - 优先选 H3,fallback H2,再 fallback 最深的 >=5 个的层级
    """
    counter = Counter()
    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            lv = heading_level(block)
            if lv is not None:
                counter[lv] += 1
    # H3 优先,然后 H2,然后挑 count>=5 中最浅的
    for candidate in (3, 2, 4):
        if counter.get(candidate, 0) >= 5:
            return candidate
    # fallback: 选 count 最多的层级,但排除 H1
    # H1 通常是文档大标题,只 1 个,选了等于把整篇文档塞成 1 个模块
    counter.pop(1, None)
    if counter:
        return counter.most_common(1)[0][0]
    print(
        "[警告] 文档没有 H2-H6 heading,模块切分无法生效。"
        "请用 --heading-level N 显式指定。当前 fallback 到 H2。",
        file=sys.stderr,
    )
    return 2


def cell_text(cell) -> str:
    return normalize("\n".join(p.text for p in cell.paragraphs))


def table_header(t: DocxTable) -> list[str]:
    if not t.rows:
        return []
    return [cell_text(c) for c in t.rows[0].cells]


def find_col(header: list[str], candidates: tuple[str, ...]) -> int | None:
    for i, h in enumerate(header):
        for cand in candidates:
            if cand in h:
                return i
    return None


def is_module_pricing_table(header: list[str]) -> bool:
    """识别"功能模块 + 费用/工作量"内嵌表。"""
    has_module = find_col(header, ("功能模块", "模块", "功能名称")) is not None
    has_money = find_col(header, ("费用", "报价", "金额", "单价", "价格")) is not None
    has_effort = find_col(header, ("工作量", "工时", "人月", "人天", "人日")) is not None
    return has_module and (has_money or has_effort)


# ---------------------------------------------------------------------------
# 抽取主逻辑
# ---------------------------------------------------------------------------


def extract_from_doc(
    doc: DocxDocumentT, module_h_level: int
) -> tuple[list[ModuleEntry], list[SectionEntry], dict[str, dict]]:
    """从 docx 抽取模块 + 章节明细 + 内嵌费用表。

    Args:
        doc: 已加载的 DocxDocument 对象 (调用方负责 load,异常处理也在调用方)
        module_h_level: 模块所在 heading 层级

    Returns:
        modules: 模块概览(每个 module heading 一项)
        sections: 章节明细(所有 paragraph)
        pricing_map: {module_title: {cost, effort}} — 内嵌表抽出的成本数据
    """
    modules: list[ModuleEntry] = []
    sections: list[SectionEntry] = []
    pricing_map: dict[str, dict] = {}

    heading_stack: list[tuple[int, str]] = []  # 当前 heading 路径
    current_module: ModuleEntry | None = None
    current_module_paragraphs: list[str] = []

    def commit_current_module():
        nonlocal current_module, current_module_paragraphs
        if current_module:
            current_module.description = " | ".join(current_module_paragraphs)[:500]
            modules.append(current_module)
        current_module = None
        current_module_paragraphs = []

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            text = normalize(block.text)
            if not text:
                continue
            lv = heading_level(block)
            if lv is not None:
                # 维护 heading 路径栈
                while heading_stack and heading_stack[-1][0] >= lv:
                    heading_stack.pop()
                heading_stack.append((lv, text))
                sections.append(SectionEntry(
                    module_title=current_module.title if current_module else "",
                    heading_level=lv, text=text,
                ))
                if lv == module_h_level:
                    commit_current_module()
                    parent = " > ".join(t for _, t in heading_stack[:-1])
                    current_module = ModuleEntry(
                        index=len(modules) + 1,
                        title=text,
                        parent_path=parent,
                        description="",
                    )
            else:
                if current_module:
                    current_module_paragraphs.append(text)
                sections.append(SectionEntry(
                    module_title=current_module.title if current_module else "",
                    heading_level=0, text=text,
                ))
        elif isinstance(block, DocxTable):
            header = table_header(block)
            if is_module_pricing_table(header):
                pricing_map.update(_extract_pricing_from_table(block, header))

    commit_current_module()

    # 把 pricing_map 回填到 modules
    for m in modules:
        if m.title in pricing_map:
            m.estimated_cost = pricing_map[m.title].get("cost", "")
            m.estimated_effort = pricing_map[m.title].get("effort", "")

    return modules, sections, pricing_map


def _extract_pricing_from_table(t: DocxTable, header: list[str]) -> dict[str, dict]:
    """从"功能模块/费用/工作量"内嵌表抽出 {模块名: {cost, effort, note}}。"""
    module_col = find_col(header, ("功能模块", "模块", "功能名称"))
    money_col = find_col(header, ("费用", "报价", "金额", "单价", "价格"))
    effort_col = find_col(header, ("工作量", "工时", "人月", "人天", "人日"))
    desc_col = find_col(header, ("备注", "说明"))
    if module_col is None:
        return {}

    # 跳过表头行,有的表"功能模块"列出现 2 次(主模块 + 子模块),取后面那个非空
    module_cols = [i for i, h in enumerate(header) if find_col([h], ("功能模块", "模块")) == 0]

    out = {}
    for row in t.rows[1:]:
        cells = [cell_text(c) for c in row.cells]
        if not cells:
            continue
        # 模块名: 优先用最后一个非空"模块"列
        module_name = ""
        for col_i in reversed(module_cols if module_cols else [module_col]):
            if col_i < len(cells) and cells[col_i].strip():
                module_name = cells[col_i].strip()
                break
        if not module_name or module_name in ("功能模块", "模块", "合计", "小计"):
            continue
        out[module_name] = {
            "cost": cells[money_col].strip() if money_col is not None and money_col < len(cells) else "",
            "effort": cells[effort_col].strip() if effort_col is not None and effort_col < len(cells) else "",
            "note": cells[desc_col].strip() if desc_col is not None and desc_col < len(cells) else "",
        }
    return out


# ---------------------------------------------------------------------------
# Excel 输出
# ---------------------------------------------------------------------------

def _set_column_widths(ws: Worksheet, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_module_sheet(ws: Worksheet, modules: list[ModuleEntry]):
    headers = ["序号", "模块名", "所属章节", "模块描述", "预估费用", "预估工作量", "备注"]
    _apply_header(ws, headers)
    for i, m in enumerate(modules, 2):
        ws.cell(row=i, column=1, value=m.index)
        ws.cell(row=i, column=2, value=m.title)
        ws.cell(row=i, column=3, value=m.parent_path)
        ws.cell(row=i, column=4, value=m.description)
        ws.cell(row=i, column=5, value=m.estimated_cost)
        ws.cell(row=i, column=6, value=m.estimated_effort)
        ws.cell(row=i, column=7, value=m.note)
    _apply_data_styles(ws, len(modules), len(headers))
    _set_column_widths(ws, [6, 28, 36, 50, 14, 14, 24])


def write_section_sheet(ws: Worksheet, sections: list[SectionEntry]):
    headers = ["所属模块", "Heading 层级", "文本"]
    _apply_header(ws, headers)
    for i, s in enumerate(sections, 2):
        ws.cell(row=i, column=1, value=s.module_title)
        ws.cell(row=i, column=2, value=s.heading_level if s.heading_level else "(段落)")
        ws.cell(row=i, column=3, value=s.text[:1000])
    _apply_data_styles(ws, len(sections), len(headers))
    _set_column_widths(ws, [22, 14, 100])


def write_readme_sheet(ws: Worksheet, source: Path, module_h_level: int, n_modules: int):
    lines = [
        ("解决方案 · 功能模块清单 · README", True),
        ("", False),
        ("用途", True),
        ("从解决方案 / 建设方案 / 升级方案 .docx 中抽取功能模块,放到 Excel 跟踪。", False),
        ("跟 PRD-aggregator 的区别: 解决方案文档以章节为主、没有 FR 编号表、可能有内嵌'功能模块 + 费用'表。", False),
        ("", False),
        ("Sheet 说明", True),
        ("• 模块概览 — 每行 = 一个功能模块,从指定 Heading 层级抽出,带上下级章节路径。", False),
        ("• 章节明细 — 文档所有 paragraph 与 heading,用于人工审阅原文。", False),
        ("• README — 本说明。", False),
        ("", False),
        ("本次抽取参数", True),
        (f"• 源文件: {source.name}", False),
        (f"• 模块 Heading 层级: H{module_h_level}", False),
        (f"• 抽出模块数: {n_modules}", False),
        ("", False),
        ("再跑命令", True),
        (f"python extract_solution.py --input {source.name} --output <out>.xlsx --heading-level {module_h_level}", False),
        ("", False),
        ("提示", True),
        ("• 如果模块数量明显偏少 / 偏多,试着改 --heading-level (常见值 2/3/4)", False),
        ("• 内嵌'功能模块 + 费用 + 工作量'表会自动识别并回填到模块概览", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = _font(bold=bold, size=12 if bold else 10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 22 if bold else 18
    ws.column_dimensions["A"].width = 100


def write_excel(modules, sections, source: Path, module_h_level: int, output: Path):
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    write_module_sheet(wb.create_sheet("模块概览"), modules)
    write_section_sheet(wb.create_sheet("章节明细"), sections)
    write_readme_sheet(wb.create_sheet("README"), source, module_h_level, len(modules))
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description="解决方案/方案文档功能模块抽取")
    ap.add_argument("--input", required=True, type=Path, help="输入 .docx 路径")
    ap.add_argument("--output", type=Path, default=None, help="输出 .xlsx (默认: 同目录同名)")
    ap.add_argument("--heading-level", type=int, default=None, help="模块所在 Heading 层级 (默认自动)")
    ap.add_argument("--inspect", action="store_true", help="只打印 Heading 层级分布")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"找不到输入: {args.input}", file=sys.stderr)
        return 2

    try:
        doc = DocxDocument(str(args.input))
    except Exception as e:
        print(f"无法打开 docx ({args.input}): {e}", file=sys.stderr)
        return 2

    if args.inspect or args.heading_level is None:
        counter = Counter()
        for block in iter_block_items(doc):
            if isinstance(block, Paragraph):
                lv = heading_level(block)
                if lv is not None:
                    counter[lv] += 1
        print(f"Heading 分布 ({args.input.name}):")
        for lv in sorted(counter):
            print(f"  H{lv}: {counter[lv]} 个")

    h_level = args.heading_level if args.heading_level is not None else detect_heading_level(doc)

    if args.inspect:
        print(f"\n推荐模块 Heading 层级: H{h_level} (用 --heading-level 覆盖)")
        return 0

    modules, sections, pricing = extract_from_doc(doc, h_level)
    print(f"模块 Heading 层级: H{h_level}")
    print(f"抽出: 模块={len(modules)} 段落={len(sections)} 内嵌报价表={len(pricing)} 行")

    output = args.output if args.output else args.input.with_suffix(".功能模块.xlsx")
    write_excel(modules, sections, args.input, h_level, output)
    print(f"已生成: {output}")

    if not modules:
        print(
            "[警告] 未识别到任何模块,可能 heading 层级估错或文档结构非标准 "
            "(用 --inspect 看 heading 分布,或 --heading-level N 显式指定)",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
