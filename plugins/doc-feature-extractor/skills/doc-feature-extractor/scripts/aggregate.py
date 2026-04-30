"""PRD 功能汇总 Excel 生成器(可复用模板)

把若干 .docx 形态的 PRD/需求确认单抽成一份带样式的 6 sheet Excel:
  Sheet 1 · 能力清单         (跨项目复用视角,主表,~6-12 行,带手填列)
  Sheet 2 · 业务模块概览     (项目内子模块跟踪,~10-20 行,带手填列)
  Sheet 3 · 需求明细         (FR-XXX 或兜底章节切段 ~50-200 行,只读)
  Sheet 4 · AI 字段定义      (各 PRD 字段表的列对齐 union)
  Sheet 5 · 文档生命周期总览 (需求确认单/签字PDF/开发PRD 三阶段是否齐备 + 状态公式)
  Sheet 6 · README

特性:
  - 异构 PRD 兜底抽取(无 FR 编号表时按章节切段)
  - 手填列保留(能力清单 + 业务模块 + 字段定义 三层独立主键,重跑不丢)
  - 能力清单层从 CAPABILITY_CATALOG 配置 + 业务模块聚合派生(N:N 关联)
  - 业务模块级聚合(从 FR 抽取结果 groupby 出来,不重读 docx)
  - openpyxl 排版 + 条件格式

【复用方式】
  1. 把这个文件复制到目标项目的 scripts/ 下
  2. 修改下方 ┏━ EDIT FOR YOUR PROJECT 区块(路径/前缀/表头关键字/兜底层级)
  3. ┗━ END EDIT 之后的代码通常不用改
  4. 跑 --dry-run 看抽取覆盖度,再正式生成,最后用 LibreOffice recalc.py 验 0 错误

依赖: python-docx, openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from docx import Document
from docx.document import Document as DocxDocument
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from _common import heading_level, iter_block_items, normalize

# =========================================================================
# ┏━ EDIT FOR YOUR PROJECT ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 改下面这一段就够了。已用 # ★ 标出最常需要调的项。
# =========================================================================

# ★ 路径(默认向上探测含 PRD-开发/ 与 PRD-需求确认/ 的目录,即项目根)
#   兼容两种存放: 项目根/scripts/aggregate.py 与 项目根/.claude/skills/prd-aggregator/scripts/aggregate.py
#   如需手动指定可改成: ROOT = Path("/path/to/project")
def _find_project_root() -> Path:
    here = Path(__file__).resolve()
    for parent in [here, *here.parents]:
        if (parent / "PRD-开发").is_dir() and (parent / "PRD-需求确认").is_dir():
            return parent
    return here.parent.parent  # legacy fallback
ROOT = _find_project_root()

# ★ 输入目录 + 输出文件
PRD_DEV_DIR = ROOT / "PRD-开发"        # 给开发的 PRD .docx
PRD_CONFIRM_DIR = ROOT / "PRD-需求确认"  # 需求确认单 .docx
SIGNED_PDF_DIR = ROOT / "已签字-PDF"     # 已签字 PDF(可选,目录不存在脚本会跳过)
ARCHIVED_DIR = ROOT / "_归档"           # 归档(可选)
OUTPUT_XLSX = ROOT / "PRD-功能汇总表.xlsx"

# ★ 文件名前缀方案:正则 group(1) 即"前缀"。例:`04_PRD_xxx.docx` → "04"
#   共享前缀(如两功能合写一份文档)用 "06-07" 形式,脚本会同时归属 06 与 07。
PREFIX_RE = re.compile(r"^(\d+(?:-\d+)?)[_\s]")

# ★ 显式列出所有前缀(决定 Sheet 3 的行顺序、用作"应该有但缺失"的检查清单)
PREFIXES = ["01-03", "04", "05", "06", "07", "08"]

# ★ 前缀 → 默认模块名(显示在 Excel 里;子模块多的前缀通过 fr_module_splitter 细分)
PREFIX_DEFAULT_MODULE = {
    "01-03": "AI 应用第一批次",
    "04": "财报解析与分析",
    "05": "投资协议 AI 辅助审查",
    "06": "BP 解析回填增强",
    "07": "296X 行业标签智能补标",
    "08": "BP 智能评估系统",
}

# ★ 能力清单 — Sheet 1 主表(跨项目复用视角,~6-12 行)
#   每个能力点对应一组 (前缀, 子模块);允许 N:N(同一子模块属于多个能力)。
#   命名要"通用"(去掉项目内部编码,如 296X 类的术语放复用注意事项里)。
#   把同一底层能力的多版本字段集扩展合并成一个能力(不要拆出独立行)。
CAPABILITY_CATALOG: list[dict] = [
    {
        "id": "C01",
        "name": "BP 智能解析",
        "default_description": (
            "从 PPT/PDF/Word 格式 BP 文件抽取结构化字段"
            "(基础信息、业务、财务、轮次等);支持多轮字段集扩展。"
        ),
        "modules": [
            ("01-03", "BP解析与访谈问题生成"),
            ("06", "BP 解析回填增强"),
        ],
    },
    {
        "id": "C02",
        "name": "基于 BP 的智能访谈问题生成",
        "default_description": "基于 BP 内容自动生成投资人访谈问题清单。",
        "modules": [
            ("01-03", "BP解析与访谈问题生成"),  # 与 C01 共享
        ],
    },
    {
        "id": "C03",
        "name": "访谈纪要智能解析",
        "default_description": "Word/音频形式访谈记录转结构化纪要(参与人、议题、要点、行动项)。",
        "modules": [
            ("01-03", "访谈纪要智能解析"),
        ],
    },
    {
        "id": "C04",
        "name": "投资协议智能解析",
        "default_description": (
            "从 Word/PDF 投资协议中抽取结构化条款字段"
            "(签约主体、金额、关键条款、对赌、回购等)。"
        ),
        "modules": [
            ("01-03", "投资协议智能解析"),
            ("05", "协议解析字段扩展"),
            ("05", "基金关键条款字段扩展"),
            ("05", "项目对赌条款日期数组实现（遗漏补充）"),
        ],
    },
    {
        "id": "C05",
        "name": "投资协议智能辅助审查",
        "default_description": "基于 LLM 对协议条款做合规检查与风险提示,在表单底部展示审查结论。",
        "modules": [
            ("05", "智能辅助审查"),
            ("05", "码值变更"),
        ],
    },
    {
        "id": "C06",
        "name": "财报智能解析",
        "default_description": "从 PDF/Word/Excel 财报中抽取财务指标(三大报表、关键比率)。",
        "modules": [
            ("04", "财报解析与分析"),
        ],
    },
    {
        "id": "C07",
        "name": "行业分类标签智能补标",
        "default_description": "基于行业知识库对历史项目缺失的行业分类标签做自动补标。",
        "modules": [
            ("07", "296X 行业标签智能补标"),
        ],
    },
    {
        "id": "C08",
        "name": "BP 智能评估",
        "default_description": "基于 BP 多维度智能打分 + 评估报告生成(含 Word 导出)。",
        "modules": [
            ("08", "业务线选择与 BP 上传"),
            ("08", "智能分析与进度展示"),
            ("08", "五维评估体系"),
            ("08", "评估报告查看"),
            ("08", "Word 报告导出"),
        ],
    },
]

# ★ 哪些前缀只有需求确认单还没 PRD —— 从 PRD_CONFIRM_DIR 抽 FR、Sheet 3 标"缺 PRD"
PREFIXES_FROM_CONFIRMATION = {"08"}

# ★ 表头关键字:用于在 .docx 表格里识别 FR 表 / AI 字段表
#   命中以下任一即为 FR 编号列:
FR_NUMBER_KEYS = ("编号", "FR编号", "FR 编号", "需求编号", "Requirement ID", "ID")
#   命中以下任一即为 FR 描述列:
FR_DESC_KEYS = ("需求描述", "描述", "需求", "Description", "Requirement")
#   AI 字段表识别: NAME 命中 + (KEY 或 TYPE) 命中
FIELD_NAME_KEYS = ("字段名称", "字段名", "信息项", "Agent输出字段", "字段", "Field Name", "Field")
FIELD_KEY_KEYS = ("字段Key", "字段 Key", "AI 返回 Key", "AI返回Key", "字段标识",
                  "Key", "表单字段 Key", "表单字段Key")
TYPE_KEYS = ("类型", "Type", "Data Type")
REQUIRED_KEYS = ("必填", "Required")
CONSTRAINT_KEYS = ("校验规则", "约束", "Constraint", "Validation")
DESC_KEYS_FIELD = ("说明", "描述", "备注", "Description", "Notes")
SCENARIO_KEYS = ("适用场景", "场景", "Scenario")

# ★ FR 编号匹配模式(章节兜底失败时也用此判断 cell 是不是真 FR 行)
FR_ID_PATTERN = re.compile(r"^FR-?\d+|^REQ-?\d+|^US-?\d+", re.IGNORECASE)

# ★ 兜底章节抽取的标题层级(没 FR 表时按这个层级切段成 <前缀>-S<n>)
#   不在表里的前缀会跳过兜底,只走 FR 表
FALLBACK_HEADING_LEVEL = {"04": 2, "06": 3, "07": 3, "08": 2}

# ★ 仅有"信息项 + 备注"这种简表的前缀(放宽 AI 字段表识别条件)
LENIENT_FIELD_PREFIXES = {"08"}

# ★ 跳过的章节关键字(版本记录/附录/Prompt 等不是功能点)
SKIP_SECTION_KEYWORDS = (
    "版本记录", "澄清记录", "附录", "Prompt", "调用OCR", "示例",
    "开放问题", "假设", "依赖", "范围外", "边缘情况",
    "Version History", "Appendix", "Examples",
)


def fr_module_splitter(prefix: str, h_module: str, fr_id: str) -> str:
    """按 FR 编号细分子模块(可选,默认 return h_module)。

    示例:01-03 PRD 内有 3 个子模块,FR 编号首数字决定:
      FR-1xx → BP 解析与访谈问题生成
      FR-2xx → 访谈纪要智能解析
      FR-3xx → 投资协议智能解析
    其他项目若 FR 编号没有这种语义,直接 return h_module 即可。
    """
    if prefix == "01-03":
        m = re.match(r"FR-?(\d)", fr_id)
        if m:
            return {
                "1": "BP解析与访谈问题生成",
                "2": "访谈纪要智能解析",
                "3": "投资协议智能解析",
            }.get(m.group(1), h_module)
    if prefix == "05":
        # 05 的 FR-3xy 中第二位 y 决定子模块(对齐字段抽取的 H1 — 注意全角括号)
        m = re.match(r"FR-?3(\d)", fr_id)
        if m:
            return {
                "1": "智能辅助审查",            # FR-31x
                "2": "智能辅助审查",            # FR-32x 仍属审查
                "3": "协议解析字段扩展",         # FR-33x
                "4": "基金关键条款字段扩展",      # FR-34x
                "5": "项目对赌条款日期数组实现（遗漏补充）",  # FR-35x — 全角括号匹配 docx H1
                "6": "码值变更",                # FR-360
            }.get(m.group(1), h_module)
    return h_module or PREFIX_DEFAULT_MODULE.get(prefix, "")


# =========================================================================
# ┗━ END EDIT  下面通常不用改 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# =========================================================================

# ---------------------------------------------------------------------------
# 排版常量
# ---------------------------------------------------------------------------

FONT_NAME = "Arial"
FONT_NAME_CN = "PingFang SC"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
MANUAL_HEADER_FILL = PatternFill("solid", fgColor="FFC000")
TITLE_FILL = PatternFill("solid", fgColor="305496")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
SUCCESS_FILL = PatternFill("solid", fgColor="E2EFDA")

PREFIX_FILLS = {
    "01-03": PatternFill("solid", fgColor="DEEBF7"),
    "04": PatternFill("solid", fgColor="E2EFDA"),
    "05": PatternFill("solid", fgColor="EAD3F2"),
    "06": PatternFill("solid", fgColor="FCE4B6"),
    "07": PatternFill("solid", fgColor="FFF2CC"),
    "08": PatternFill("solid", fgColor="E7E6E6"),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


# ---------------------------------------------------------------------------
# 数据类
# ---------------------------------------------------------------------------

MANUAL_MODULE_COLS = ("关键能力", "开发状态", "责任人", "优先级", "计划交付", "备注")
MANUAL_FIELD_COLS = ("回填实现状态", "测试通过率", "备注")
MANUAL_CAPABILITY_COLS = (
    "能力描述", "复用度", "成熟度", "已应用项目", "复用注意事项", "维护责任人",
)


@dataclass
class FRRow:
    prefix: str
    module: str
    fr_id: str
    description: str
    source_doc: str
    doc_type: str


@dataclass
class FieldRow:
    prefix: str
    module: str
    field_name: str
    field_key: str
    field_type: str
    required: str
    constraint: str
    description: str
    scenario: str
    source_doc: str
    source_section: str
    manual: dict = field(default_factory=dict)


@dataclass
class LifecycleRow:
    prefix: str
    module: str
    confirmation: str
    signed_pdf: str
    dev_prd: str
    archived: str


@dataclass
class ModuleRow:
    """业务模块级聚合行 — Sheet 2(项目内子模块跟踪)。"""

    prefix: str
    module: str
    doc_type: str
    fr_count: int
    field_count: int
    key_capability: str
    source_doc: str
    capability_ids: list[str] = field(default_factory=list)
    manual: dict = field(default_factory=dict)


@dataclass
class CapabilityRow:
    """能力清单行 — Sheet 1 主表(跨项目复用视角)。"""

    id: str
    name: str
    default_description: str
    modules: list[tuple[str, str]]
    source_docs: list[str] = field(default_factory=list)
    manual: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# docx 工具
# ---------------------------------------------------------------------------


_HEADING_NUMBER_RE = re.compile(r"^\d+(\.\d+)*[.\s、]\s*")


def strip_heading_number(text: str) -> str:
    """去掉 '2. BP解析' / '3.1 功能概述' 这类章节序号前缀。"""
    return _HEADING_NUMBER_RE.sub("", text).strip()


def file_prefix(name: str) -> str | None:
    m = PREFIX_RE.match(name)
    return m.group(1) if m else None


def table_header(t: DocxTable) -> list[str]:
    if not t.rows:
        return []
    return [normalize(c.text) for c in t.rows[0].cells]


def find_col(header: list[str], candidates: tuple[str, ...]) -> int | None:
    for i, h in enumerate(header):
        if h in candidates:
            return i
    return None


# ---------------------------------------------------------------------------
# FR 抽取
# ---------------------------------------------------------------------------


def is_fr_table(header: list[str]) -> bool:
    return (
        find_col(header, FR_NUMBER_KEYS) is not None
        and find_col(header, FR_DESC_KEYS) is not None
    )


def is_field_table(header: list[str], lenient: bool = False) -> bool:
    name_idx = find_col(header, FIELD_NAME_KEYS)
    if name_idx is None:
        return False
    if lenient:
        return True
    has_key = find_col(header, FIELD_KEY_KEYS) is not None
    has_type = find_col(header, TYPE_KEYS) is not None
    return has_key or has_type


def extract_fr_rows_from_doc(
    doc: DocxDocument, prefix: str, source_doc: str, doc_type: str
) -> list[FRRow]:
    """优先抽 FR 表;若无,按 H2/H3 章节切段兜底。"""
    rows: list[FRRow] = []
    current_h1 = ""
    current_h2 = ""
    fr_seen = False

    # 第一遍: 找 FR 表
    for block in iter_block_items(doc):
        if isinstance(block, DocxTable) and is_fr_table(table_header(block)):
            header = table_header(block)
            num_idx = find_col(header, FR_NUMBER_KEYS)
            desc_idx = find_col(header, FR_DESC_KEYS)
            if num_idx is None or desc_idx is None:
                continue
            module = current_h1 or current_h2 or PREFIX_DEFAULT_MODULE.get(prefix, "")
            for r in block.rows[1:]:
                cells = [normalize(c.text) for c in r.cells]
                fr_id = cells[num_idx] if num_idx < len(cells) else ""
                desc = cells[desc_idx] if desc_idx < len(cells) else ""
                if fr_id and FR_ID_PATTERN.match(fr_id):
                    rows.append(
                        FRRow(prefix, fr_module_splitter(prefix, module, fr_id),
                              fr_id, desc, source_doc, doc_type)
                    )
                    fr_seen = True
        elif isinstance(block, Paragraph):
            level = heading_level(block)
            if level == 1:
                current_h1 = strip_heading_number(normalize(block.text))
                current_h2 = ""
            elif level == 2:
                current_h2 = strip_heading_number(normalize(block.text))

    if fr_seen:
        return rows

    # 兜底: 按章节切段
    rows = _section_fallback_fr(doc, prefix, source_doc, doc_type)
    return rows


def _section_fallback_fr(
    doc: DocxDocument, prefix: str, source_doc: str, doc_type: str
) -> list[FRRow]:
    """按章节切段生成 <prefix>-S<n>。"""
    rows: list[FRRow] = []
    current_h1 = ""
    target_level = FALLBACK_HEADING_LEVEL.get(prefix)
    if target_level is None:
        return rows  # 该前缀未配置兜底
    seq = 0
    pending: tuple[str, str] | None = None  # (section_name, first_para)
    has_para = False

    def flush():
        nonlocal pending, seq, has_para
        if pending is None:
            return
        seq += 1
        section_name, first_para = pending
        fr_id = f"{prefix}-S{seq:02d}"
        module = _module_from_section(prefix, current_h1, section_name)
        desc = first_para or section_name
        rows.append(FRRow(prefix, module, fr_id, desc, source_doc, doc_type))
        pending = None
        has_para = False

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            level = heading_level(block)
            text = normalize(block.text)
            section_text = strip_heading_number(text) if level else text
            if level == 1:
                current_h1 = section_text
            if level == target_level and section_text:
                if _is_skippable_section(section_text):
                    continue
                flush()
                pending = (section_text, "")
                has_para = False
            elif pending and not has_para and text:
                pending = (pending[0], text[:300])
                has_para = True
        elif isinstance(block, DocxTable) and pending and not has_para:
            # 章节里没有正文段落,用首个表格的第一行作描述
            header = table_header(block)
            if header and len(block.rows) > 1:
                sample = [normalize(c.text) for c in block.rows[1].cells]
                pending = (pending[0], " / ".join(s for s in sample if s)[:300])
                has_para = True
    flush()
    return rows


def _is_multi_module_prefix(prefix: str) -> bool:
    """前缀是否是多子模块:
       01-03 / 05 ✓ — 一份 PRD 内多个 H1 子模块
       08      ✓ — 一份需求确认单内多个 H2 子模块
       04/06/07 ✗ — 一份 PRD 一个模块
    """
    if prefix in PREFIXES_FROM_CONFIRMATION:
        return True
    return prefix not in FALLBACK_HEADING_LEVEL


def _module_from_section(prefix: str, h1: str, section_name: str) -> str:
    """章节兜底时的模块名 — 必须跟字段抽取保持一致:
       08(仅需求确认): H2 章节就是子模块 → 用 section_name
       单模块(04/06/07): 永远用 PREFIX_DEFAULT_MODULE,忽略 docx 实际 H1 文本
       多模块(理论上不走兜底,但万一走): 用 H1
    """
    if prefix in PREFIXES_FROM_CONFIRMATION:
        return section_name or PREFIX_DEFAULT_MODULE.get(prefix, "")
    if not _is_multi_module_prefix(prefix):
        return PREFIX_DEFAULT_MODULE.get(prefix, h1)
    return h1 or PREFIX_DEFAULT_MODULE.get(prefix, "")


def _is_skippable_section(text: str) -> bool:
    for kw in SKIP_SECTION_KEYWORDS:
        if kw in text:
            return True
    return False


# ---------------------------------------------------------------------------
# AI 字段抽取
# ---------------------------------------------------------------------------


def extract_field_rows_from_doc(
    doc: DocxDocument, prefix: str, source_doc: str
) -> list[FieldRow]:
    rows: list[FieldRow] = []
    current_section = ""
    current_h1 = ""
    current_h2 = ""
    lenient = prefix in LENIENT_FIELD_PREFIXES

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            level = heading_level(block)
            if level:
                text = strip_heading_number(normalize(block.text))
                if level == 1:
                    current_h1 = text
                    current_h2 = ""
                elif level == 2:
                    current_h2 = text
                current_section = text
            continue

        header = table_header(block)
        if not is_field_table(header, lenient=lenient):
            continue
        if _is_skippable_section(current_section):
            continue

        name_idx = find_col(header, FIELD_NAME_KEYS)
        key_idx = find_col(header, FIELD_KEY_KEYS)
        type_idx = find_col(header, TYPE_KEYS)
        req_idx = find_col(header, REQUIRED_KEYS)
        cons_idx = find_col(header, CONSTRAINT_KEYS)
        desc_idx = find_col(header, DESC_KEYS_FIELD)
        scen_idx = find_col(header, SCENARIO_KEYS)

        # 模块归属(必须跟 FR 抽取保持一致,否则 aggregate_module_rows 会拆成两行):
        #   仅需求确认单(08)— H2 是子模块
        #   多子模块(01-03, 05)— H1 是子模块
        #   单模块(04/06/07)— 强制用 PREFIX_DEFAULT_MODULE
        if prefix in PREFIXES_FROM_CONFIRMATION:
            module = current_h2 or current_h1 or PREFIX_DEFAULT_MODULE.get(prefix, "")
        elif _is_multi_module_prefix(prefix):
            module = current_h1 or PREFIX_DEFAULT_MODULE.get(prefix, "")
        else:
            module = PREFIX_DEFAULT_MODULE.get(prefix, "")

        for r in block.rows[1:]:
            cells = [normalize(c.text) for c in r.cells]

            def get(idx: int | None) -> str:
                return cells[idx] if idx is not None and idx < len(cells) else ""

            name_raw = get(name_idx)
            if not name_raw:
                continue
            field_name, parsed_key = _split_name_key(name_raw)
            field_key = get(key_idx) or parsed_key
            rows.append(
                FieldRow(
                    prefix=prefix,
                    module=module,
                    field_name=field_name,
                    field_key=field_key,
                    field_type=get(type_idx),
                    required=get(req_idx),
                    constraint=get(cons_idx),
                    description=get(desc_idx),
                    scenario=get(scen_idx),
                    source_doc=source_doc,
                    source_section=current_section,
                )
            )
    return rows


_NAME_KEY_RE = re.compile(r"^(.+?)[（(]\s*([A-Za-z][\w\d]*)\s*[）)]\s*$")


def _split_name_key(raw: str) -> tuple[str, str]:
    """'企业全称（name）' -> ('企业全称', 'name')。"""
    m = _NAME_KEY_RE.match(raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw, ""


# ---------------------------------------------------------------------------
# 生命周期表
# ---------------------------------------------------------------------------


def build_lifecycle_rows() -> list[LifecycleRow]:
    confirmations = _list_files(PRD_CONFIRM_DIR, ".docx")
    signed = _list_files(SIGNED_PDF_DIR, ".pdf")
    devs = _list_files(PRD_DEV_DIR, ".docx")
    archived = _list_files(ARCHIVED_DIR)

    rows: list[LifecycleRow] = []
    for prefix in PREFIXES:
        confirm_files = _match(confirmations, prefix)
        signed_files = _match(signed, prefix)
        dev_files = _match(devs, prefix)
        arch_files = _match(archived, prefix)

        rows.append(
            LifecycleRow(
                prefix=prefix,
                module=PREFIX_DEFAULT_MODULE.get(prefix, ""),
                confirmation=", ".join(f.name for f in confirm_files) or "✗",
                signed_pdf=", ".join(f.name for f in signed_files) or "✗",
                dev_prd=", ".join(f.name for f in dev_files) or "✗",
                archived=", ".join(f.name for f in arch_files),
            )
        )
    return rows


def _list_files(d: Path, ext: str | None = None) -> list[Path]:
    if not d.exists():
        return []
    out = []
    for f in sorted(d.iterdir()):
        if f.name.startswith("."):
            continue
        if ext and f.suffix.lower() != ext:
            continue
        out.append(f)
    return out


def _match(files: list[Path], prefix: str) -> list[Path]:
    """`06-07` 类共享文件需同时归属 06 与 07。"""
    out = []
    for f in files:
        fp = file_prefix(f.name)
        if fp == prefix:
            out.append(f)
            continue
        # 共享前缀: 06-07 命中 06 / 07
        if fp and "-" in fp and prefix != fp:
            parts = fp.split("-")
            if len(parts) == 2 and prefix in parts:
                out.append(f)
    return out


# ---------------------------------------------------------------------------
# 抽取入口
# ---------------------------------------------------------------------------


def extract_all() -> tuple[list[FRRow], list[FieldRow], list[LifecycleRow]]:
    fr_rows: list[FRRow] = []
    field_rows: list[FieldRow] = []

    prefixes_from_dev = [p for p in PREFIXES if p not in PREFIXES_FROM_CONFIRMATION]

    # 有 PRD 的前缀: 从 PRD-开发/ 抽
    for prefix in prefixes_from_dev:
        for f in _list_files(PRD_DEV_DIR, ".docx"):
            if file_prefix(f.name) != prefix:
                continue
            doc = Document(str(f))
            fr_rows.extend(extract_fr_rows_from_doc(doc, prefix, f.name, "开发PRD"))
            field_rows.extend(extract_field_rows_from_doc(doc, prefix, f.name))

    # 仅需求确认单的前缀: 从 PRD-需求确认/ 抽,文档类型标"仅需求确认单(缺PRD)"
    for prefix in PREFIXES_FROM_CONFIRMATION:
        for f in _list_files(PRD_CONFIRM_DIR, ".docx"):
            if file_prefix(f.name) != prefix:
                continue
            doc = Document(str(f))
            fr_rows.extend(extract_fr_rows_from_doc(doc, prefix, f.name, "仅需求确认单(缺PRD)"))
            field_rows.extend(extract_field_rows_from_doc(doc, prefix, f.name))

    lifecycle_rows = build_lifecycle_rows()
    return fr_rows, field_rows, lifecycle_rows


# ---------------------------------------------------------------------------
# 手填列保留
# ---------------------------------------------------------------------------


def load_existing_manual(path: Path) -> tuple[dict, dict, dict]:
    """读旧 Excel 的手填列,返回 (capability_manual, module_manual, field_manual)。

    - capability_manual: {capability_id: {col: value}}
    - module_manual: {(prefix, module): {col: value}}
    - field_manual: {(prefix, module, field_name): {col: value}}

    兼容多种旧版本:
      v3 (能力清单 + 业务模块概览): 都读
      v2 (业务模块概览): 没能力清单 sheet,只读模块手填
      v1 (功能需求清单, FR 级): 旧 Sheet 1 是 80 行 FR,把同一模块下任一 FR 的手填值聚合(取第一个非空)
    """
    capability_manual: dict[str, dict[str, object]] = {}
    module_manual: dict[tuple[str, str], dict[str, object]] = {}
    field_manual: dict[tuple[str, str, str], dict[str, object]] = {}
    if not path.exists():
        return capability_manual, module_manual, field_manual
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[警告] 无法读取旧 Excel ({e}),跳过手填列保留", file=sys.stderr)
        return capability_manual, module_manual, field_manual

    # ── 能力清单 sheet(v3+) ─────────────────────────────
    if "能力清单" in wb.sheetnames:
        ws = wb["能力清单"]
        header = [c.value for c in ws[2]] if ws.max_row >= 2 else []
        idx_map = {h: i for i, h in enumerate(header) if h}
        if "能力 ID" in idx_map or "能力ID" in idx_map:
            id_col = idx_map.get("能力 ID", idx_map.get("能力ID"))
            for row in ws.iter_rows(min_row=3, values_only=True):
                if id_col is None or id_col >= len(row):
                    continue
                cap_id = row[id_col]
                if not cap_id:
                    continue
                capability_manual[str(cap_id)] = {
                    col: row[idx_map[col]]
                    for col in MANUAL_CAPABILITY_COLS
                    if col in idx_map and idx_map[col] < len(row) and row[idx_map[col]] is not None
                }

    # ── Sheet 业务模块概览: 优先 v2 业务模块概览,兜底 v1 功能需求清单 ─────────────
    if "业务模块概览" in wb.sheetnames:
        ws = wb["业务模块概览"]
        header = [c.value for c in ws[2]] if ws.max_row >= 2 else []
        idx_map = {h: i for i, h in enumerate(header) if h}
        if all(k in idx_map for k in ("功能前缀", "子模块")):
            for row in ws.iter_rows(min_row=3, values_only=True):
                p = row[idx_map["功能前缀"]] if idx_map["功能前缀"] < len(row) else None
                m = row[idx_map["子模块"]] if idx_map["子模块"] < len(row) else None
                if not p or not m:
                    continue
                key = (str(p), str(m))
                module_manual[key] = {
                    col: row[idx_map[col]]
                    for col in MANUAL_MODULE_COLS
                    if col in idx_map and idx_map[col] < len(row) and row[idx_map[col]] is not None
                }
    elif "功能需求清单" in wb.sheetnames:
        # v1 迁移: 同一 (prefix, module) 下取第一个非空手填值
        ws = wb["功能需求清单"]
        header = [c.value for c in ws[2]] if ws.max_row >= 2 else []
        idx_map = {h: i for i, h in enumerate(header) if h}
        v1_cols = ("开发状态", "责任人", "优先级", "计划交付", "备注")
        if all(k in idx_map for k in ("功能前缀", "模块名称")):
            multi_value_modules: set[tuple[str, str]] = set()
            for row in ws.iter_rows(min_row=3, values_only=True):
                p = row[idx_map["功能前缀"]] if idx_map["功能前缀"] < len(row) else None
                m = row[idx_map["模块名称"]] if idx_map["模块名称"] < len(row) else None
                if not p or not m:
                    continue
                key = (str(p), str(m))
                bucket = module_manual.setdefault(key, {})
                for col in v1_cols:
                    if col not in idx_map or idx_map[col] >= len(row):
                        continue
                    val = row[idx_map[col]]
                    if val is None or str(val).strip() == "":
                        continue
                    if col not in bucket:
                        bucket[col] = val
                    elif str(bucket[col]) != str(val):
                        # 多个 FR 在同一模块下手填了不同值
                        multi_value_modules.add(key)
            if multi_value_modules:
                for key in multi_value_modules:
                    note_prefix = "[迁移自多个 FR 的手填值,请人工核对]"
                    existing = str(module_manual[key].get("备注", "") or "")
                    if note_prefix not in existing:
                        module_manual[key]["备注"] = (note_prefix + " " + existing).strip()

    # ── Sheet 2: AI 字段定义(沿用旧逻辑,主键 (prefix, module, field_name)) ─────────
    if "AI 字段定义" in wb.sheetnames:
        ws = wb["AI 字段定义"]
        header = [c.value for c in ws[2]] if ws.max_row >= 2 else []
        idx_map = {h: i for i, h in enumerate(header) if h}
        if all(k in idx_map for k in ("功能前缀", "模块名称", "字段名称")):
            for row in ws.iter_rows(min_row=3, values_only=True):
                p = row[idx_map["功能前缀"]]
                m = row[idx_map["模块名称"]]
                n = row[idx_map["字段名称"]]
                if not n:
                    continue
                key = (str(p or ""), str(m or ""), str(n))
                field_manual[key] = {
                    col: row[idx_map[col]]
                    for col in MANUAL_FIELD_COLS
                    if col in idx_map and idx_map[col] < len(row) and row[idx_map[col]] is not None
                }
    return capability_manual, module_manual, field_manual


def merge_manual(
    capability_rows: list[CapabilityRow],
    module_rows: list[ModuleRow],
    field_rows: list[FieldRow],
    capability_manual: dict,
    module_manual: dict,
    field_manual: dict,
) -> tuple[list[CapabilityRow], list[ModuleRow], list[FieldRow]]:
    """合并旧手填列。「关键能力」「能力描述」特殊处理: 旧值非空时不覆盖自动填的摘要。"""
    # ── 能力清单 ──────────────────────────────────
    seen_caps: set[str] = set()
    for r in capability_rows:
        seen_caps.add(r.id)
        if r.id in capability_manual:
            old = capability_manual[r.id]
            for col in MANUAL_CAPABILITY_COLS:
                v = old.get(col)
                if v is None or str(v).strip() == "":
                    continue
                r.manual[col] = v
    # 旧条目: catalog 里删了的能力 → 保留行,加标注
    for cap_id, manual in capability_manual.items():
        if cap_id in seen_caps:
            continue
        if not any(manual.values()):
            continue
        ghost = CapabilityRow(
            id=cap_id,
            name=str(manual.get("能力描述", "") or "[已从 catalog 移除]")[:40],
            default_description="[已从 catalog 移除,但手填列保留]",
            modules=[],
            source_docs=[],
            manual=dict(manual),
        )
        existing_note = str(ghost.manual.get("复用注意事项", "") or "")
        ghost.manual["复用注意事项"] = (
            f"[已从 catalog 移除] {existing_note}".strip()
        )
        capability_rows.append(ghost)

    # ── 业务模块 ──────────────────────────────────
    seen_modules: set[tuple[str, str]] = set()
    for r in module_rows:
        key = (r.prefix, r.module)
        seen_modules.add(key)
        if key in module_manual:
            old = module_manual[key]
            for col in MANUAL_MODULE_COLS:
                v = old.get(col)
                if v is None or str(v).strip() == "":
                    continue
                if col == "关键能力":
                    # 用户改过的关键能力优先于自动摘要
                    r.key_capability = str(v)
                else:
                    r.manual[col] = v

    # 旧条目: 子模块没了 → 保留行并标记
    for key, manual in module_manual.items():
        if key in seen_modules:
            continue
        if not any(manual.values()):
            continue
        prefix, module = key
        ghost = ModuleRow(
            prefix=prefix,
            module=f"{module}(已删)",
            doc_type="—",
            fr_count=0,
            field_count=0,
            key_capability=str(manual.get("关键能力", "") or "[已从源 PRD 删除,但手填列保留]"),
            source_doc="—",
            manual={k: v for k, v in manual.items() if k != "关键能力"},
        )
        ghost.manual["备注"] = (
            f"[已从源 PRD 删除] {ghost.manual.get('备注', '') or ''}".strip()
        )
        module_rows.append(ghost)

    # AI 字段表 (沿用旧逻辑)
    seen_field: set[tuple[str, str, str]] = set()
    for r in field_rows:
        key3 = (r.prefix, r.module, r.field_name)
        if key3 in field_manual:
            r.manual = dict(field_manual[key3])
        seen_field.add(key3)
    for key3, manual in field_manual.items():
        if key3 in seen_field:
            continue
        if not any(manual.values()):
            continue
        prefix, module, name = key3
        ghost = FieldRow(
            prefix=prefix,
            module=f"{module}(已删)",
            field_name=name,
            field_key="",
            field_type="",
            required="",
            constraint="",
            description="[已从源 PRD 删除,但手填列保留]",
            scenario="",
            source_doc="—",
            source_section="—",
            manual=dict(manual),
        )
        ghost.manual["备注"] = (
            f"[已从源 PRD 删除] {ghost.manual.get('备注', '') or ''}".strip()
        )
        field_rows.append(ghost)

    return capability_rows, module_rows, field_rows


def _build_module_to_capabilities(
    catalog: list[dict],
) -> dict[tuple[str, str], list[str]]:
    """反向索引: (prefix, module) → 该模块归属的所有能力 ID(允许 N:N)。"""
    out: dict[tuple[str, str], list[str]] = {}
    for cap in catalog:
        cap_id = cap.get("id", "")
        for mod_key in cap.get("modules", []):
            out.setdefault(tuple(mod_key), []).append(cap_id)
    return out


def aggregate_module_rows(
    fr_rows: list[FRRow],
    field_rows: list[FieldRow],
    catalog: list[dict] | None = None,
) -> list[ModuleRow]:
    """从 fr_rows 和 field_rows 聚合出业务模块级 ModuleRow。

    模块边界已经在 fr_rows.module / field_rows.module 里(由 fr_module_splitter 和章节兜底决定)。
    单纯 groupby 即可,不需要重读 docx。

    如果传了 catalog,顺便从 CAPABILITY_CATALOG 反向算每行的 capability_ids。
    """
    fr_by_key: dict[tuple[str, str], list[FRRow]] = {}
    for r in fr_rows:
        fr_by_key.setdefault((r.prefix, r.module), []).append(r)

    field_count_by_key: dict[tuple[str, str], int] = {}
    for r in field_rows:
        field_count_by_key[(r.prefix, r.module)] = field_count_by_key.get((r.prefix, r.module), 0) + 1

    module_to_caps = _build_module_to_capabilities(catalog or [])

    rows: list[ModuleRow] = []
    seen: set[tuple[str, str]] = set()

    # 用 fr_rows 里的顺序保证稳定排序(后续 write 函数会再排一遍)
    ordered_keys: list[tuple[str, str]] = []
    for r in fr_rows:
        k = (r.prefix, r.module)
        if k not in seen:
            seen.add(k)
            ordered_keys.append(k)
    # 字段表里独有的模块也要加(虽然实际中很少见)
    for r in field_rows:
        k = (r.prefix, r.module)
        if k not in seen:
            seen.add(k)
            ordered_keys.append(k)

    for key in ordered_keys:
        prefix, module = key
        frs = fr_by_key.get(key, [])
        first_desc = frs[0].description if frs else ""
        # 关键能力: 取首个 FR 描述前 60 字
        key_capability = first_desc[:60] if first_desc else ""
        # 文档类型 & 来源文档: 取该模块的第一个 FR
        doc_type = frs[0].doc_type if frs else ""
        source_doc = frs[0].source_doc if frs else ""
        rows.append(
            ModuleRow(
                prefix=prefix,
                module=module,
                doc_type=doc_type,
                fr_count=len(frs),
                field_count=field_count_by_key.get(key, 0),
                key_capability=key_capability,
                source_doc=source_doc,
                capability_ids=list(module_to_caps.get(key, [])),
            )
        )
    return rows


def aggregate_capability_rows(
    catalog: list[dict],
    module_rows: list[ModuleRow],
) -> list[CapabilityRow]:
    """从 CAPABILITY_CATALOG 配置 + 已聚合的 module_rows 派生出能力清单行。

    纯聚合,不读 docx。来源文档列从 module_rows 反查(去重)。
    """
    module_to_source: dict[tuple[str, str], str] = {
        (m.prefix, m.module): m.source_doc for m in module_rows
    }
    rows: list[CapabilityRow] = []
    for cap in catalog:
        modules = [tuple(m) for m in cap.get("modules", [])]
        source_docs: list[str] = []
        for mk in modules:
            doc = module_to_source.get(mk, "")
            if doc and doc not in source_docs:
                source_docs.append(doc)
        rows.append(
            CapabilityRow(
                id=cap["id"],
                name=cap["name"],
                default_description=cap.get("default_description", ""),
                modules=list(modules),  # type: ignore[arg-type]
                source_docs=source_docs,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Excel 写入
# ---------------------------------------------------------------------------


def write_excel(
    capability_rows: list[CapabilityRow],
    module_rows: list[ModuleRow],
    fr_rows: list[FRRow],
    field_rows: list[FieldRow],
    lifecycle_rows: list[LifecycleRow],
    path: Path,
) -> None:
    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    sheet_cap = wb.create_sheet("能力清单")
    sheet_module = wb.create_sheet("业务模块概览")
    sheet_fr = wb.create_sheet("需求明细")
    sheet_field = wb.create_sheet("AI 字段定义")
    sheet_life = wb.create_sheet("文档生命周期总览")
    sheet_readme = wb.create_sheet("README")

    write_capability_sheet(sheet_cap, capability_rows)
    write_module_sheet(sheet_module, module_rows)
    write_fr_sheet(sheet_fr, fr_rows)
    write_field_sheet(sheet_field, field_rows)
    write_lifecycle_sheet(sheet_life, lifecycle_rows)
    write_readme_sheet(sheet_readme)

    wb.save(path)


def _write_title(ws: Worksheet, title: str, n_cols: int) -> None:
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1).font = _font(bold=True, color="FFFFFF", size=13)
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _write_header(ws: Worksheet, headers: list[str], manual_cols: set[str]) -> None:
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        is_manual = h in manual_cols
        c.font = _font(bold=True, color="000000" if is_manual else "FFFFFF", size=11)
        c.fill = MANUAL_HEADER_FILL if is_manual else HEADER_FILL
        c.alignment = CENTER_CENTER
        c.border = BORDER
    ws.row_dimensions[2].height = 28


def _apply_data_styles(
    ws: Worksheet,
    n_rows: int,
    n_cols: int,
    long_text_cols: set[int],
    prefix_col: int | None = None,
    rows_data: list | None = None,
) -> None:
    for r in range(3, 3 + n_rows):
        zebra = (r - 2) % 2 == 0
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _font()
            cell.alignment = WRAP_CENTER if c in long_text_cols else CENTER_CENTER
            cell.border = BORDER
            if zebra:
                cell.fill = ALT_ROW_FILL
        if prefix_col and rows_data:
            data_idx = r - 3
            if data_idx < len(rows_data):
                pfx = getattr(rows_data[data_idx], "prefix", None)
                if pfx in PREFIX_FILLS:
                    ws.cell(row=r, column=prefix_col).fill = PREFIX_FILLS[pfx]
        ws.row_dimensions[r].height = 22


def _set_column_widths(
    ws: Worksheet, widths: list[int]
) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_capability_sheet(ws: Worksheet, rows: list[CapabilityRow]) -> None:
    """Sheet 1 · 能力清单(跨项目复用视角,~6-12 行,带手填列)。"""
    headers = [
        "能力 ID", "能力名称", "能力描述", "关联业务模块", "文档来源",
        "复用度", "成熟度", "已应用项目", "复用注意事项", "维护责任人",
    ]
    n_cols = len(headers)
    _write_title(ws, f"能力清单 — 最后生成: {datetime.now():%Y-%m-%d %H:%M}", n_cols)
    _write_header(ws, headers, set(MANUAL_CAPABILITY_COLS))

    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r.id)
        ws.cell(row=i, column=2, value=r.name)
        # 能力描述: 半手填,旧值优先,否则用 default_description
        desc_value = r.manual.get("能力描述", "") or r.default_description
        ws.cell(row=i, column=3, value=desc_value)
        # 关联业务模块: 自动,逗号分隔 "(前缀) 子模块"
        modules_text = ", ".join(f"{p} {m}" for p, m in r.modules)
        ws.cell(row=i, column=4, value=modules_text)
        # 文档来源: 自动,逗号分隔
        ws.cell(row=i, column=5, value=", ".join(r.source_docs))
        # 5 个纯手填列
        ws.cell(row=i, column=6, value=r.manual.get("复用度", ""))
        ws.cell(row=i, column=7, value=r.manual.get("成熟度", ""))
        ws.cell(row=i, column=8, value=r.manual.get("已应用项目", ""))
        ws.cell(row=i, column=9, value=r.manual.get("复用注意事项", ""))
        ws.cell(row=i, column=10, value=r.manual.get("维护责任人", ""))

    _apply_data_styles(
        ws,
        n_rows=len(rows),
        n_cols=n_cols,
        long_text_cols={3, 4, 5, 8, 9},
        prefix_col=None,
        rows_data=None,
    )
    _set_column_widths(
        ws, [8, 22, 50, 40, 30, 10, 12, 25, 40, 12]
    )

    ws.freeze_panes = "C3"
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{2 + len(rows)}"

    if not rows:
        return
    last_row = 2 + len(rows)
    rng = f"A3:{get_column_letter(n_cols)}{last_row}"
    # 成熟度=已上线 → 浅绿;成熟度=POC → 浅黄
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=['$G3="已上线"'], stopIfTrue=False, fill=SUCCESS_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=['$G3="POC"'],
            stopIfTrue=False,
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    # 复用度=低 → 单元格浅红(只对 F 列)
    ws.conditional_formatting.add(
        f"F3:F{last_row}",
        FormulaRule(formula=['$F3="低"'], stopIfTrue=False, fill=WARNING_FILL),
    )


def write_module_sheet(ws: Worksheet, rows: list[ModuleRow]) -> None:
    """Sheet 2 · 业务模块概览(项目内跟踪,~16 行,带手填列)。"""
    headers = [
        "功能前缀", "子模块", "所属能力", "文档类型", "FR数", "字段数", "来源文档",
        "关键能力", "开发状态", "责任人", "优先级", "计划交付", "备注",
    ]
    n_cols = len(headers)
    _write_title(ws, f"业务模块概览 — 最后生成: {datetime.now():%Y-%m-%d %H:%M}", n_cols)
    _write_header(ws, headers, set(MANUAL_MODULE_COLS))

    rows_sorted = sorted(
        rows,
        key=lambda r: (PREFIXES.index(r.prefix) if r.prefix in PREFIXES else 99, r.module),
    )
    for i, r in enumerate(rows_sorted, start=3):
        ws.cell(row=i, column=1, value=r.prefix)
        ws.cell(row=i, column=2, value=r.module)
        ws.cell(row=i, column=3, value=", ".join(r.capability_ids) if r.capability_ids else "")
        ws.cell(row=i, column=4, value=r.doc_type)
        ws.cell(row=i, column=5, value=r.fr_count)
        ws.cell(row=i, column=6, value=r.field_count)
        ws.cell(row=i, column=7, value=r.source_doc)
        ws.cell(row=i, column=8, value=r.key_capability)  # 关键能力(半手填)
        ws.cell(row=i, column=9, value=r.manual.get("开发状态", ""))
        ws.cell(row=i, column=10, value=r.manual.get("责任人", ""))
        ws.cell(row=i, column=11, value=r.manual.get("优先级", ""))
        ws.cell(row=i, column=12, value=r.manual.get("计划交付", ""))
        ws.cell(row=i, column=13, value=r.manual.get("备注", ""))

    _apply_data_styles(
        ws,
        n_rows=len(rows_sorted),
        n_cols=n_cols,
        long_text_cols={2, 3, 7, 8, 13},
        prefix_col=1,
        rows_data=rows_sorted,
    )

    # 未归类(没有任何能力 ID)→ 序号格(A 列)加浅红覆盖前缀色,提示该模块未挂能力
    for idx, r in enumerate(rows_sorted, start=3):
        if not r.capability_ids:
            ws.cell(row=idx, column=1).fill = WARNING_FILL
    _set_column_widths(
        ws, [10, 30, 14, 18, 8, 8, 36, 50, 12, 12, 10, 14, 30]
    )

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{2 + len(rows_sorted)}"

    # 条件格式: 开发状态 列 I(原 H,前面插了「所属能力」)
    last_row = 2 + len(rows_sorted)
    rng = f"A3:{get_column_letter(n_cols)}{last_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=['$I3="已完成"'], stopIfTrue=False, fill=SUCCESS_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=['$I3="阻塞"'], stopIfTrue=False, fill=WARNING_FILL),
    )


def write_fr_sheet(ws: Worksheet, rows: list[FRRow]) -> None:
    """Sheet 2 · 需求明细(只读,FR/章节兜底级 ~80 行,无手填列)。"""
    headers = [
        "功能前缀", "子模块", "FR编号", "需求描述", "来源文档", "文档类型",
    ]
    n_cols = len(headers)
    _write_title(ws, f"需求明细(只读) — 最后生成: {datetime.now():%Y-%m-%d %H:%M}", n_cols)
    _write_header(ws, headers, set())

    rows_sorted = sorted(
        rows, key=lambda r: (PREFIXES.index(r.prefix) if r.prefix in PREFIXES else 99, r.fr_id)
    )
    for i, r in enumerate(rows_sorted, start=3):
        ws.cell(row=i, column=1, value=r.prefix)
        ws.cell(row=i, column=2, value=r.module)
        ws.cell(row=i, column=3, value=r.fr_id)
        ws.cell(row=i, column=4, value=r.description)
        ws.cell(row=i, column=5, value=r.source_doc)
        ws.cell(row=i, column=6, value=r.doc_type)

    _apply_data_styles(
        ws,
        n_rows=len(rows_sorted),
        n_cols=n_cols,
        long_text_cols={2, 4, 5},
        prefix_col=1,
        rows_data=rows_sorted,
    )
    _set_column_widths(ws, [10, 28, 14, 60, 40, 22])

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{2 + len(rows_sorted)}"


def write_field_sheet(ws: Worksheet, rows: list[FieldRow]) -> None:
    headers = [
        "功能前缀", "模块名称", "字段名称", "字段Key", "类型", "必填", "约束",
        "说明", "适用场景", "来源章节", "来源文档",
        "回填实现状态", "测试通过率", "备注",
    ]
    n_cols = len(headers)
    _write_title(ws, f"AI 字段定义 — 最后生成: {datetime.now():%Y-%m-%d %H:%M}", n_cols)
    _write_header(ws, headers, set(MANUAL_FIELD_COLS))

    rows_sorted = sorted(rows, key=lambda r: (PREFIXES.index(r.prefix) if r.prefix in PREFIXES else 99, r.module, r.field_name))
    for i, r in enumerate(rows_sorted, start=3):
        ws.cell(row=i, column=1, value=r.prefix)
        ws.cell(row=i, column=2, value=r.module)
        ws.cell(row=i, column=3, value=r.field_name)
        ws.cell(row=i, column=4, value=r.field_key)
        ws.cell(row=i, column=5, value=r.field_type)
        ws.cell(row=i, column=6, value=r.required)
        ws.cell(row=i, column=7, value=r.constraint)
        ws.cell(row=i, column=8, value=r.description)
        ws.cell(row=i, column=9, value=r.scenario)
        ws.cell(row=i, column=10, value=r.source_section)
        ws.cell(row=i, column=11, value=r.source_doc)
        for j, col in enumerate(MANUAL_FIELD_COLS, start=12):
            ws.cell(row=i, column=j, value=r.manual.get(col, ""))

    _apply_data_styles(
        ws,
        n_rows=len(rows_sorted),
        n_cols=n_cols,
        long_text_cols={2, 7, 8, 9, 10, 14},
        prefix_col=1,
        rows_data=rows_sorted,
    )
    _set_column_widths(
        ws, [10, 24, 22, 22, 14, 8, 24, 50, 24, 30, 36, 14, 14, 30]
    )

    # 测试通过率列百分比格式
    pct_col = headers.index("测试通过率") + 1
    for r in range(3, 3 + len(rows_sorted)):
        ws.cell(row=r, column=pct_col).number_format = "0.0%"

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{2 + len(rows_sorted)}"


def write_lifecycle_sheet(ws: Worksheet, rows: list[LifecycleRow]) -> None:
    headers = [
        "功能前缀", "模块名称", "需求确认单", "已签字PDF", "开发PRD", "归档版本",
        "状态备注",
    ]
    n_cols = len(headers)
    _write_title(ws, f"文档生命周期总览 — 最后生成: {datetime.now():%Y-%m-%d %H:%M}", n_cols)
    _write_header(ws, headers, set())

    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r.prefix)
        ws.cell(row=i, column=2, value=r.module)
        ws.cell(row=i, column=3, value=r.confirmation)
        ws.cell(row=i, column=4, value=r.signed_pdf)
        ws.cell(row=i, column=5, value=r.dev_prd)
        ws.cell(row=i, column=6, value=r.archived)
        # 公式: 缺什么自动判定
        ws.cell(
            row=i,
            column=7,
            value=(
                f'=IF(E{i}="✗","缺 PRD",'
                f'IF(D{i}="✗","缺签字PDF",'
                f'IF(C{i}="✗","缺需求确认单","齐备")))'
            ),
        )

    _apply_data_styles(
        ws,
        n_rows=len(rows),
        n_cols=n_cols,
        long_text_cols={3, 4, 5, 6, 7},
        prefix_col=1,
        rows_data=rows,
    )
    _set_column_widths(
        ws, [10, 28, 56, 56, 50, 36, 18]
    )

    last_row = 2 + len(rows)

    # 统计汇总行
    summary_row = last_row + 1
    ws.cell(row=summary_row, column=1, value="汇总").font = _font(bold=True, size=11)
    ws.cell(row=summary_row, column=2, value="共 / 齐备 / 缺 PRD / 缺签字PDF / 缺需求确认单").font = _font(bold=True)
    ws.cell(
        row=summary_row,
        column=7,
        value=(
            f'=COUNTA(A3:A{last_row})&" / "&'
            f'COUNTIF(G3:G{last_row},"齐备")&" / "&'
            f'COUNTIF(G3:G{last_row},"缺 PRD")&" / "&'
            f'COUNTIF(G3:G{last_row},"缺签字PDF")&" / "&'
            f'COUNTIF(G3:G{last_row},"缺需求确认单")'
        ),
    )
    ws.cell(row=summary_row, column=7).font = _font(bold=True)
    for c in range(1, n_cols + 1):
        ws.cell(row=summary_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=summary_row, column=c).border = BORDER
        ws.cell(row=summary_row, column=c).alignment = CENTER_CENTER

    # 条件格式: 状态备注含"缺" → 整行浅红
    rng = f"A3:{get_column_letter(n_cols)}{last_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("缺",$G3))'],
            stopIfTrue=False,
            fill=WARNING_FILL,
        ),
    )

    ws.freeze_panes = "C3"


def write_readme_sheet(ws: Worksheet) -> None:
    fallback_prefixes = ", ".join(sorted(FALLBACK_HEADING_LEVEL.keys()))
    confirm_prefixes = ", ".join(sorted(PREFIXES_FROM_CONFIRMATION)) or "(无)"
    lines = [
        ("PRD 功能汇总表 · README", True),
        ("", False),
        ("用途", True),
        ("这是一份 AI 能力库的拆解视图 — 把核心 AI 技术拆成最小可独立定价、独立评估、独立复用的业务单元。", False),
        ("• 客户问「这个能力值多少钱、能不能给下个项目复用」 → 看「能力清单」", False),
        ("• PM 问「这个模块谁在做、到哪了」 → 看「业务模块概览」", False),
        ("• 测试问「AI 抽得准不准」 → 看「AI 字段定义」", False),
        ("• 配套追踪所有 PRD/需求确认单 的功能点与生命周期(签字 PDF/开发 PRD 是否齐备)。", False),
        ("", False),
        ("Sheet 说明", True),
        ("• 能力清单 — 跨项目复用视角,每行 = 一个可复用的功能能力点(~6-12 行)。打开默认看到这张。手填列:能力描述 / 复用度 / 成熟度 / 已应用项目 / 复用注意事项 / 维护责任人。", False),
        ("• 业务模块概览 — 项目内子模块跟踪,每行 = 这个项目的一个业务子模块(~16 行)。手填列:关键能力 / 开发状态 / 责任人 / 优先级 / 计划交付 / 备注。「所属能力」列指向上一张 sheet。", False),
        ("• 需求明细 — 只读,每行 = 一个 FR 或章节切段(~80 行),业务表 FR 数 / 字段数对应这里的细节。", False),
        ("• AI 字段定义 — 各 PRD 字段定义表的并集,列对齐 union,缺的留空。", False),
        ("• 文档生命周期总览 — 每个前缀一行,三阶段是否齐备(需求确认单/已签字PDF/开发PRD),状态备注是公式自动算的。", False),
        ("• README — 本说明。", False),
        ("", False),
        ("两层语义的差别(关键)", True),
        ("• 能力清单 = 「我们沉淀了什么,下个客户能复用什么」:跨项目视角,粒度粗,删去项目内部术语。", False),
        ("• 业务模块概览 = 「这个项目里干了什么活」:项目跟踪视角,粒度跟着 PRD 章节走。", False),
        ("• 一个能力可以横跨多个业务模块(如 BP 解析 = 01-03 BP 解析 + 06 字段扩展);一个业务模块也可能挂多个能力(如 BP 解析与访谈问题生成 同时属于 C01 和 C02)。", False),
        ("", False),
        ("手填列(橙黄表头) vs 自动列(深蓝表头)", True),
        ("• 自动列由脚本生成,重跑会被覆盖。", False),
        ("• 能力清单手填列按 能力 ID 主键合并;业务模块概览按 (前缀, 子模块) 主键合并;AI 字段定义按 (前缀, 模块, 字段名) 主键合并。三层互不干扰。", False),
        ("• 「关键能力」「能力描述」首次自动填,你改写过之后视作手填,重跑不覆盖。", False),
        ("• 子模块/字段/能力被源删除时,会保留行并在备注前加 [已从源 PRD 删除] 或 [已从 catalog 移除] 标记,不直接抹掉。", False),
        ("• 业务模块概览 A 列(功能前缀)如果是浅红填充,代表该模块未挂到任何能力点 — 检查 CAPABILITY_CATALOG 是否需要新增条目。", False),
        ("", False),
        ("重新生成命令", True),
        ("python scripts/aggregate.py --dry-run  # 干跑统计", False),
        ("python scripts/aggregate.py            # 正式生成", False),
        ("", False),
        ("本项目配置(脚本顶部 EDIT 区块)", True),
        (f"• 前缀清单: {', '.join(PREFIXES)}", False),
        (f"• 兜底章节切段的前缀: {fallback_prefixes or '(无)'}", False),
        (f"• 仅需求确认单的前缀(标记缺 PRD): {confirm_prefixes}", False),
        ("• 若新 PRD 表头不命中 FIELD_NAME_KEYS / FIELD_KEY_KEYS / TYPE_KEYS,在脚本顶部扩这几个常量元组。", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = _font(bold=bold, size=12 if bold else 10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 22 if bold else 18
    ws.column_dimensions["A"].width = 120


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description="PRD 功能汇总 Excel 生成器")
    ap.add_argument(
        "--input",
        type=Path,
        default=None,
        help=(
            "项目根目录(含 PRD-开发/ 与 PRD-需求确认/)。"
            "默认从脚本所在路径向上探测;指定后可跨项目调用,无需改脚本顶部 EDIT 区块。"
        ),
    )
    ap.add_argument("--dry-run", action="store_true", help="只统计抽取行数,不写文件")
    ap.add_argument(
        "--output",
        "--output-xlsx",
        dest="output_xlsx",
        type=Path,
        default=None,
        help=(
            "覆盖输出 xlsx 路径(默认用脚本顶部 OUTPUT_XLSX 或 <input>/PRD-功能汇总表.xlsx)。"
            "适合 benchmark / 沙箱场景: 不改脚本就能把产物落到指定目录。"
        ),
    )
    args = ap.parse_args()

    if args.input is not None:
        global ROOT, PRD_DEV_DIR, PRD_CONFIRM_DIR, SIGNED_PDF_DIR, ARCHIVED_DIR, OUTPUT_XLSX
        ROOT = args.input.resolve()
        if not (ROOT / "PRD-开发").is_dir() and not (ROOT / "PRD-需求确认").is_dir():
            print(
                f"[警告] {ROOT} 下既无 PRD-开发/ 也无 PRD-需求确认/,可能不是合法项目根",
                file=sys.stderr,
            )
        PRD_DEV_DIR = ROOT / "PRD-开发"
        PRD_CONFIRM_DIR = ROOT / "PRD-需求确认"
        SIGNED_PDF_DIR = ROOT / "已签字-PDF"
        ARCHIVED_DIR = ROOT / "_归档"
        OUTPUT_XLSX = ROOT / "PRD-功能汇总表.xlsx"

    output_path = args.output_xlsx if args.output_xlsx else OUTPUT_XLSX

    fr_rows, field_rows, lifecycle_rows = extract_all()

    if args.dry_run:
        print("=== Dry-run 抽取统计 ===")
        from collections import Counter

        fr_counter = Counter(r.prefix for r in fr_rows)
        field_counter = Counter(r.prefix for r in field_rows)
        for p in PREFIXES:
            print(
                f"  [{p}] FR={fr_counter.get(p, 0):>3}  字段={field_counter.get(p, 0):>3}"
            )
        print(f"  生命周期行数 = {len(lifecycle_rows)}")
        for r in lifecycle_rows:
            confirm_status = "✓" if r.confirmation != "✗" else "✗"
            signed_status = "✓" if r.signed_pdf != "✗" else "✗"
            dev_status = "✓" if r.dev_prd != "✗" else "✗"
            print(
                f"    {r.prefix}: 需求确认{confirm_status} 签字PDF{signed_status} 开发PRD{dev_status}"
            )
        return 0

    module_rows = aggregate_module_rows(fr_rows, field_rows, CAPABILITY_CATALOG)
    capability_rows = aggregate_capability_rows(CAPABILITY_CATALOG, module_rows)
    capability_manual, module_manual, field_manual = load_existing_manual(output_path)
    capability_rows, module_rows, field_rows = merge_manual(
        capability_rows, module_rows, field_rows,
        capability_manual, module_manual, field_manual,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(capability_rows, module_rows, fr_rows, field_rows, lifecycle_rows, output_path)
    print(f"已生成: {output_path}")
    print(
        f"  能力清单 {len(capability_rows)} 行,"
        f"业务模块 {len(module_rows)} 行,需求明细 {len(fr_rows)} 行,"
        f"AI 字段 {len(field_rows)} 行,生命周期 {len(lifecycle_rows)} 行"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
