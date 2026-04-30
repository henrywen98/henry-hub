"""测 extract_solution.py (Mode B) 的端到端行为。"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

import extract_solution

SCRIPT = Path(__file__).resolve().parent.parent / "scripts" / "extract_solution.py"


def _run(argv: list[str]) -> subprocess.CompletedProcess:
    """跑 extract_solution.py 子进程,捕获 stdout/stderr/exit code。"""
    return subprocess.run(
        [sys.executable, str(SCRIPT), *argv],
        capture_output=True,
        text=True,
    )


def test_happy_path_synthetic_solution(synthetic_solution_docx: Path, tmp_path: Path):
    out_xlsx = tmp_path / "out.xlsx"
    proc = _run(["--input", str(synthetic_solution_docx), "--output", str(out_xlsx)])
    assert proc.returncode == 0, proc.stderr
    assert out_xlsx.exists()

    wb = load_workbook(out_xlsx, data_only=True)
    assert "模块概览" in wb.sheetnames
    assert "章节明细" in wb.sheetnames
    assert "README" in wb.sheetnames

    ws = wb["模块概览"]
    headers = [c.value for c in ws[1]]
    expected_eval_cols = ["复用度", "成熟度", "已应用项目", "维护责任人", "备注"]
    for col in expected_eval_cols:
        assert col in headers, f"模块概览 sheet 缺评估列 {col}"

    # 5 个 H3 模块 → 5 行数据 + 1 行表头
    assert ws.max_row == 6
    titles = [ws.cell(row=r, column=2).value for r in range(2, 7)]
    assert "项目智能分析" in titles
    assert "BP 文件解析" in titles


def test_pricing_table_backfill(synthetic_solution_docx: Path, tmp_path: Path):
    """内嵌 (功能模块/费用/工作量) 表应该回填到模块概览的 预估费用 / 预估工作量 列。"""
    out_xlsx = tmp_path / "out.xlsx"
    _run(["--input", str(synthetic_solution_docx), "--output", str(out_xlsx)])

    wb = load_workbook(out_xlsx, data_only=True)
    ws = wb["模块概览"]
    # row 2 应该是 "项目智能分析",fixture 给的费用是 "1.5 万",工作量 "5 人天"
    cost_col = next(i for i, c in enumerate(ws[1], 1) if c.value == "预估费用")
    effort_col = next(i for i, c in enumerate(ws[1], 1) if c.value == "预估工作量")
    found_pricing = False
    for r in range(2, ws.max_row + 1):
        cost = ws.cell(row=r, column=cost_col).value
        effort = ws.cell(row=r, column=effort_col).value
        if cost and effort:
            found_pricing = True
            break
    assert found_pricing, "内嵌报价表未回填"


def test_zero_modules_returns_exit_1(synthetic_solution_no_modules_docx: Path, tmp_path: Path):
    """无 heading 的 docx → 0 模块 → exit 1 + stderr 警告 (Tier 1.2 契约)。"""
    out_xlsx = tmp_path / "out.xlsx"
    proc = _run(["--input", str(synthetic_solution_no_modules_docx), "--output", str(out_xlsx)])
    assert proc.returncode == 1, f"应 exit 1 (0 模块场景), got {proc.returncode}"
    assert "未识别到任何模块" in proc.stderr


def test_corrupt_docx_returns_exit_2(corrupt_docx: Path, tmp_path: Path):
    """0 字节假 docx → exit 2 + 友好错误 (Tier 1.1 契约)。"""
    out_xlsx = tmp_path / "out.xlsx"
    proc = _run(["--input", str(corrupt_docx), "--output", str(out_xlsx)])
    assert proc.returncode == 2
    assert "无法打开 docx" in proc.stderr


def test_rerun_preserves_manual_columns(synthetic_solution_docx: Path, tmp_path: Path):
    """T4 contract 4: 跑一遍 → 在 xlsx 写手填值 → 重跑 → 手填值必须保留。"""
    out_xlsx = tmp_path / "out.xlsx"

    # First run
    proc1 = _run(["--input", str(synthetic_solution_docx), "--output", str(out_xlsx)])
    assert proc1.returncode == 0

    # Seed manual values into row 2 (col 7 复用度, col 8 成熟度, col 11 备注)
    wb = load_workbook(out_xlsx)
    ws = wb["模块概览"]
    ws.cell(row=2, column=7, value="高")
    ws.cell(row=2, column=8, value="GA")
    ws.cell(row=2, column=11, value="测试用例 - 重跑必须保留")
    wb.save(str(out_xlsx))

    # Rerun
    proc2 = _run(["--input", str(synthetic_solution_docx), "--output", str(out_xlsx)])
    assert proc2.returncode == 0
    assert "手填保留" in proc2.stderr

    # Verify
    wb = load_workbook(out_xlsx, data_only=True)
    ws = wb["模块概览"]
    assert ws.cell(row=2, column=7).value == "高"
    assert ws.cell(row=2, column=8).value == "GA"
    assert ws.cell(row=2, column=11).value == "测试用例 - 重跑必须保留"


def test_inspect_mode_lists_heading_distribution(synthetic_solution_docx: Path):
    """--inspect 应输出 heading 分布并返回 0,不写文件。"""
    proc = _run(["--input", str(synthetic_solution_docx), "--inspect"])
    assert proc.returncode == 0
    assert "Heading 分布" in proc.stdout
    assert "推荐模块 Heading 层级" in proc.stdout


def test_detect_heading_level_excludes_h1(synthetic_solution_docx: Path):
    """T2.4 fallback: detect_heading_level 不应返回 H1。
    本 fixture 主路径会选 H3 (count=5),H1 排除规则在 fallback 才生效。
    这里覆盖主路径,fallback 行为由 _common 侧测。
    """
    from docx import Document

    doc = Document(str(synthetic_solution_docx))
    level = extract_solution.detect_heading_level(doc)
    assert level == 3
