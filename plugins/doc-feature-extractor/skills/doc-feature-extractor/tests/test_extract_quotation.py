"""测 extract_quotation.py (Mode C) 的端到端行为。"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

import extract_quotation

SCRIPT = Path(__file__).resolve().parent.parent / "scripts" / "extract_quotation.py"


def _run(argv: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(SCRIPT), *argv],
        capture_output=True,
        text=True,
    )


def test_happy_path_synthetic_quotation(synthetic_quotation_xlsx: Path, tmp_path: Path):
    out_xlsx = tmp_path / "out.xlsx"
    proc = _run(["--input", str(synthetic_quotation_xlsx), "--output", str(out_xlsx)])
    assert proc.returncode == 0, proc.stderr

    wb = load_workbook(out_xlsx, data_only=True)
    assert "模块报价" in wb.sheetnames
    assert "README" in wb.sheetnames

    ws = wb["模块报价"]
    headers = [c.value for c in ws[1]]
    for col in ("复用度", "成熟度", "已应用项目", "维护责任人", "备注"):
        assert col in headers


def test_parent_value_inheritance(synthetic_quotation_xlsx: Path, tmp_path: Path):
    """业务范围 列只在第一行写,后面行应该继承父值 (extract_quotation 招牌行为)。"""
    out_xlsx = tmp_path / "out.xlsx"
    _run(["--input", str(synthetic_quotation_xlsx), "--output", str(out_xlsx)])

    wb = load_workbook(out_xlsx, data_only=True)
    ws = wb["模块报价"]
    scope_col = next(i for i, c in enumerate(ws[1], 1) if c.value == "业务范围")
    # fixture 第一行: 投资管理。后面 2 行应该继承 投资管理 (而不是空)
    scopes_for_first_sheet = []
    for r in range(2, ws.max_row + 1):
        ws_name = ws.cell(row=r, column=2).value
        if ws_name == "客户A-报价清单":
            scopes_for_first_sheet.append(ws.cell(row=r, column=scope_col).value)
    # 应该有至少 3 行 (PDF / PPT / 条款) 且都是 "投资管理"
    assert len(scopes_for_first_sheet) >= 3
    assert all(s == "投资管理" for s in scopes_for_first_sheet), scopes_for_first_sheet


def test_total_row_skipped(synthetic_quotation_xlsx: Path, tmp_path: Path):
    """合计 / 小计 行应该被跳过,不进入输出。"""
    out_xlsx = tmp_path / "out.xlsx"
    _run(["--input", str(synthetic_quotation_xlsx), "--output", str(out_xlsx)])

    wb = load_workbook(out_xlsx, data_only=True)
    ws = wb["模块报价"]
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                assert "合计" not in str(val), f"合计行未被过滤: row {r} col {c}"


def test_col_patterns_disambiguation(
    synthetic_quotation_ambiguous_headers_xlsx: Path, tmp_path: Path
):
    """T2.1 contract: header [功能模块, 功能说明, 备注说明] 不能让 note 抢 desc。"""
    rows = extract_quotation.extract_from_xlsx(
        synthetic_quotation_ambiguous_headers_xlsx, sheet_filter=None
    )
    assert len(rows) == 1
    r = rows[0]
    assert r.module == "模块1"
    assert r.description == "这是功能描述"  # desc 命中"功能说明"
    assert r.note == "这是备注内容"           # note 命中"备注说明"


def test_no_header_returns_exit_1(synthetic_no_header_xlsx: Path, tmp_path: Path):
    out_xlsx = tmp_path / "out.xlsx"
    proc = _run(["--input", str(synthetic_no_header_xlsx), "--output", str(out_xlsx)])
    assert proc.returncode == 1
    assert "未识别到表头" in proc.stderr


def test_corrupt_xlsx_returns_exit_1(tmp_path: Path):
    """伪 xlsx (非 zip) → load_workbook 抛异常 → 友好提示 + exit 1。"""
    bogus = tmp_path / "fake.xlsx"
    bogus.write_bytes(b"not a real xlsx")
    out_xlsx = tmp_path / "out.xlsx"
    proc = _run(["--input", str(bogus), "--output", str(out_xlsx)])
    assert proc.returncode == 1
    assert "无法打开 xlsx" in proc.stderr


def test_rerun_preserves_manual_with_hostile_override(
    synthetic_quotation_xlsx: Path, tmp_path: Path
):
    """T4 contract 4 + 敌对场景: 用户改了 备注 (本来有源备注 'A'),重跑必须保留用户值。"""
    out_xlsx = tmp_path / "out.xlsx"

    # First run
    proc1 = _run(["--input", str(synthetic_quotation_xlsx), "--output", str(out_xlsx)])
    assert proc1.returncode == 0

    # Find row that has 模块=合同审查 (源 fixture 给了 备注="源备注 A"),
    # 用户在那行覆盖 备注 + 加 复用度 + 维护责任人
    wb = load_workbook(out_xlsx)
    ws = wb["模块报价"]
    headers = [c.value for c in ws[1]]
    note_col = headers.index("备注") + 1
    reuse_col = headers.index("复用度") + 1
    maint_col = headers.index("维护责任人") + 1
    module_col = headers.index("模块") + 1

    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=module_col).value == "合同审查":
            target_row = r
            break
    assert target_row is not None
    # 验证源 备注 是 "源备注 A"
    assert ws.cell(row=target_row, column=note_col).value == "源备注 A"

    ws.cell(row=target_row, column=note_col, value="HENRY 评估覆盖")
    ws.cell(row=target_row, column=reuse_col, value="高 / 已用 2")
    ws.cell(row=target_row, column=maint_col, value="Henry")
    wb.save(str(out_xlsx))

    # Rerun
    proc2 = _run(["--input", str(synthetic_quotation_xlsx), "--output", str(out_xlsx)])
    assert proc2.returncode == 0
    assert "手填保留" in proc2.stderr

    wb = load_workbook(out_xlsx, data_only=True)
    ws = wb["模块报价"]
    # row 顺序可能变;按 模块=合同审查 重新定位
    found = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=module_col).value == "合同审查":
            assert ws.cell(row=r, column=note_col).value == "HENRY 评估覆盖"
            assert ws.cell(row=r, column=reuse_col).value == "高 / 已用 2"
            assert ws.cell(row=r, column=maint_col).value == "Henry"
            found = True
            break
    assert found
