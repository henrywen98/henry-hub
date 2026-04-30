"""测 _common.py 的纯结构工具与 load_manual_columns。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

import _common


# ---------------------------------------------------------------------------
# 结构工具
# ---------------------------------------------------------------------------


def test_normalize_collapses_whitespace_and_strips():
    assert _common.normalize("  hello   world  ") == "hello world"
    assert _common.normalize("\t\nfoo\n bar  ") == "foo bar"
    assert _common.normalize("") == ""
    assert _common.normalize(None) == ""  # type: ignore[arg-type]


def test_heading_level_matches_variants():
    """提示词 T2.3: 正则要兼容 "Heading 1" / "Heading10" / "heading 2"。"""

    class _MockStyle:
        def __init__(self, name):
            self.name = name

    class _MockParagraph:
        def __init__(self, style_name):
            self.style = _MockStyle(style_name)

    assert _common.heading_level(_MockParagraph("Heading 1")) == 1
    assert _common.heading_level(_MockParagraph("Heading10")) == 10  # 无空格
    assert _common.heading_level(_MockParagraph("heading 2")) == 2  # 小写
    assert _common.heading_level(_MockParagraph("Heading  3")) == 3  # 多空格
    assert _common.heading_level(_MockParagraph("Normal")) is None
    assert _common.heading_level(_MockParagraph("")) is None


# ---------------------------------------------------------------------------
# load_manual_columns
# ---------------------------------------------------------------------------


def _make_xlsx(tmp_path: Path, sheet_name: str, headers: list[str], rows: list[list]) -> Path:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    out = tmp_path / "fixture.xlsx"
    wb.save(out)
    return out


def test_load_manual_columns_returns_empty_when_path_missing(tmp_path: Path):
    nonexistent = tmp_path / "does_not_exist.xlsx"
    assert _common.load_manual_columns(
        nonexistent, "any", ("k",), ("v",)
    ) == {}


def test_load_manual_columns_returns_empty_when_sheet_missing(tmp_path: Path):
    path = _make_xlsx(tmp_path, "wrong_sheet", ["a"], [["1"]])
    assert _common.load_manual_columns(
        path, "right_sheet", ("a",), ("b",)
    ) == {}


def test_load_manual_columns_warns_on_missing_key_column(tmp_path: Path, capsys):
    path = _make_xlsx(tmp_path, "test", ["a", "b"], [["1", "2"]])
    out = _common.load_manual_columns(path, "test", ("nonexistent_key",), ("b",))
    assert out == {}
    captured = capsys.readouterr()
    assert "缺主键列" in captured.err


def test_load_manual_columns_happy_path(tmp_path: Path):
    path = _make_xlsx(
        tmp_path,
        "模块概览",
        ["模块名", "所属章节", "复用度", "成熟度", "备注"],
        [
            ["模块A", "升级 > X", "高", "GA", "首跑数据"],
            ["模块B", "升级 > Y", "", "Pilot", ""],  # 部分填写
            ["模块C", "升级 > Z", "", "", ""],         # 全空 → 不应进 dict
        ],
    )
    out = _common.load_manual_columns(
        path,
        "模块概览",
        key_cols=("模块名", "所属章节"),
        manual_cols=("复用度", "成熟度", "备注"),
    )
    assert ("模块A", "升级 > X") in out
    assert out[("模块A", "升级 > X")] == {"复用度": "高", "成熟度": "GA", "备注": "首跑数据"}
    assert ("模块B", "升级 > Y") in out
    assert out[("模块B", "升级 > Y")] == {"成熟度": "Pilot"}  # 只保留非空
    assert ("模块C", "升级 > Z") not in out  # 全空跳过


def test_load_manual_columns_handles_unreadable_xlsx(tmp_path: Path, capsys):
    bogus = tmp_path / "fake.xlsx"
    bogus.write_bytes(b"not a real xlsx")
    out = _common.load_manual_columns(bogus, "any", ("k",), ("v",))
    assert out == {}
    captured = capsys.readouterr()
    assert "无法读取旧 Excel" in captured.err


# ---------------------------------------------------------------------------
# 排版 helper (smoke)
# ---------------------------------------------------------------------------


def test_apply_header_marks_manual_cols_with_orange(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    _common._apply_header(ws, ["a", "b", "c", "d"], manual_col_indices=[3, 4])
    out = tmp_path / "t.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws2 = wb2.active
    # col 1, 2 → HEADER_FILL (深蓝 1F4E78)
    assert ws2.cell(row=1, column=1).fill.start_color.rgb.endswith("1F4E78")
    assert ws2.cell(row=1, column=2).fill.start_color.rgb.endswith("1F4E78")
    # col 3, 4 → MANUAL_HEADER_FILL (橙 FFC000)
    assert ws2.cell(row=1, column=3).fill.start_color.rgb.endswith("FFC000")
    assert ws2.cell(row=1, column=4).fill.start_color.rgb.endswith("FFC000")
