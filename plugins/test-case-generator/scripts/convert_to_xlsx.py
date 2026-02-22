#!/usr/bin/env python3
"""Convert test case JSON to formatted Excel (.xlsx) file.

Usage: python convert_to_xlsx.py <input.json> <output.xlsx>

All styling is driven by the JSON style block — no hardcoded formatting.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def hex_to_argb(hex_color):
    return 'FF' + hex_color.lstrip('#')


def make_font(cfg, **overrides):
    kw = {
        'name': cfg.get('font_name', '微软雅黑'),
        'size': cfg.get('font_size', 11),
        'bold': cfg.get('bold', False),
    }
    if 'font_color' in cfg:
        kw['color'] = hex_to_argb(cfg['font_color'])
    kw.update(overrides)
    return Font(**kw)


def make_fill(cfg):
    if 'fill_color' in cfg:
        return PatternFill('solid', fgColor=hex_to_argb(cfg['fill_color']))
    return PatternFill(fill_type=None)


def make_alignment(cfg):
    a = cfg.get('alignment', {})
    return Alignment(
        horizontal=a.get('horizontal', 'left'),
        vertical=a.get('vertical', 'top'),
        wrap_text=a.get('wrap_text', True)
    )


def make_border(cfg):
    if cfg.get('border') == 'thin':
        s = Side(style='thin')
        return Border(left=s, right=s, top=s, bottom=s)
    return Border()


def apply_style(cell, cfg, **font_overrides):
    cell.font = make_font(cfg, **font_overrides)
    cell.fill = make_fill(cfg)
    cell.alignment = make_alignment(cfg)
    cell.border = make_border(cfg)


def estimate_row_height(text, col_width, font_size=11):
    if not text:
        return None
    lines = str(text).split('\n')
    total_lines = 0
    for line in lines:
        chars_per_line = max(1, int(col_width * 1.5))
        total_lines += max(1, -(-len(line) // chars_per_line))  # ceil division
    height = total_lines * (font_size + 3)
    return max(15.6, min(409.5, height))


def convert(json_path, output_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    style = data['style']
    meta = data['meta']
    columns = style['header_row']['columns']
    col_count = len(columns)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = meta.get('module_name', 'Sheet1')

    # Pre-load style configs needed across multiple row sections
    sum_cfg = style['summary_bar']
    dat_cfg = style['data_row']

    # --- Row 1: Title Bar ---
    title_cfg = style['title_bar']
    for ci in range(1, col_count + 1):
        cell = ws.cell(row=1, column=ci)
        if ci == 1:
            cell.value = title_cfg['text']
        apply_style(cell, title_cfg)
    # Title bar result labels from summary_bar config
    if col_count >= 12:
        result_labels = sum_cfg.get('result_labels', ['Pass', 'Fail', 'Blocked'])
        ws.cell(row=1, column=9).value = '测试结果'
        for i, rl in enumerate(result_labels):
            ws.cell(row=1, column=10 + i).value = rl
    # M-Q: empty, lighter style from data_row config
    for ci in range(13, col_count + 1):
        cell = ws.cell(row=1, column=ci)
        apply_style(cell, dat_cfg)
    ws.row_dimensions[1].height = 19

    # --- Row 2: Summary Bar ---
    labels = sum_cfg.get('labels', [])
    for ci, label in enumerate(labels, 1):
        cell = ws.cell(row=2, column=ci)
        cell.value = label if label else None
        if ci <= 7:
            apply_style(cell, sum_cfg)
        else:
            apply_style(cell, dat_cfg)
    # Result count formulas
    result_labels = sum_cfg.get('result_labels', ['Pass', 'Fail', 'Blocked'])
    for i, label in enumerate(result_labels):
        ci = 10 + i  # J, K, L
        cell = ws.cell(row=2, column=ci)
        exec_col_idx = next((ci for ci, c in enumerate(columns, 1) if c['key'] == 'exec_result'), 14)
        exec_col = get_column_letter(exec_col_idx)
        cell.value = f'=COUNTIF({exec_col}:{exec_col},"{label}")'
        apply_style(cell, dat_cfg)
    ws.row_dimensions[2].height = 15

    # --- Row 3: Info Bar ---
    info_cfg = style['info_bar']
    info_text = f"模块名称：{meta['module_name']}\n设计人员：\n{meta.get('designer', '')}\n设计时间：{meta.get('design_date', '')}"
    merge_cols = info_cfg.get('merge_cols', 'A:C')
    merge_start, merge_end = merge_cols.split(':')
    ws.merge_cells(f'{merge_start}3:{merge_end}3')
    cell = ws.cell(row=3, column=1)
    cell.value = info_text
    cell.font = make_font(info_cfg)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = make_border({'border': 'thin'})
    for ci in range(1, col_count + 1):
        ws.cell(row=3, column=ci).border = make_border({'border': 'thin'})
    ws.row_dimensions[3].height = 60

    # --- Row 4: Header Row ---
    hdr_cfg = style['header_row']
    for ci, col_def in enumerate(columns, 1):
        cell = ws.cell(row=4, column=ci)
        cell.value = col_def['label']
        apply_style(cell, hdr_cfg)

    ws.row_dimensions[4].height = 15

    # --- Column widths ---
    for ci, col_def in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_def['width']

    # --- Data rows ---
    sec_cfg = style['section_row']
    keys = [c['key'] for c in columns]
    current_row = 5
    section_boundaries = []

    for section in data.get('sections', []):
        # Section separator row
        _ = current_row  # section start tracked by section_boundaries
        cell = ws.cell(row=current_row, column=1)
        cell.value = section['title']
        apply_style(cell, sec_cfg)
        for ci in range(2, col_count + 1):
            c = ws.cell(row=current_row, column=ci)
            c.border = make_border(sec_cfg)
            c.fill = make_fill(sec_cfg)
        ws.row_dimensions[current_row].height = 18.6
        current_row += 1

        # Test case rows
        data_start = current_row
        for tc in section.get('test_cases', []):
            max_height = 15.6
            for ci, key in enumerate(keys, 1):
                cell = ws.cell(row=current_row, column=ci)
                val = tc.get(key, '')
                cell.value = val if val else None
                apply_style(cell, dat_cfg)
                # Estimate height from steps/expected columns
                if key in ('steps', 'expected') and val:
                    col_w = columns[ci-1]['width']
                    h = estimate_row_height(val, col_w)
                    if h and h > max_height:
                        max_height = h
            ws.row_dimensions[current_row].height = max_height
            current_row += 1

        section_boundaries.append((data_start, current_row - 1))

    # --- Merge rules ---
    merge_keys = data.get('merge_rules', {})
    for key, rule in merge_keys.items():
        if rule != 'same_value':
            continue
        col_idx = None
        for ci, col_def in enumerate(columns, 1):
            if col_def['key'] == key:
                col_idx = ci
                break
        if col_idx is None:
            continue

        col_letter = get_column_letter(col_idx)
        for (start_r, end_r) in section_boundaries:
            merge_start = start_r
            for r in range(start_r + 1, end_r + 2):  # +2 to process last group
                curr_val = ws.cell(row=r, column=col_idx).value if r <= end_r else None
                prev_val = ws.cell(row=merge_start, column=col_idx).value
                if r > end_r or curr_val != prev_val or curr_val is None:
                    if r - 1 > merge_start and prev_val is not None:
                        ws.merge_cells(f'{col_letter}{merge_start}:{col_letter}{r-1}')
                    merge_start = r

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <input.json> <output.xlsx>")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
