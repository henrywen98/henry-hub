import subprocess, json, os, sys

def test_convert():
    script = os.path.join(os.path.dirname(__file__), 'convert_to_xlsx.py')
    example = os.path.join(os.path.dirname(__file__), '..', 'skills', 'test-case-generator', 'references', 'example-output.json')
    output = '/tmp/test_output.xlsx'

    if os.path.exists(output):
        os.remove(output)

    result = subprocess.run([sys.executable, script, example, output], capture_output=True, text=True)
    assert result.returncode == 0, f"Script failed: {result.stderr}"
    assert os.path.exists(output), "Output file not created"

    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active

    # Row 1: title bar
    assert ws['A1'].value == '功能测试用例'
    assert ws['A1'].font.name == '微软雅黑'
    assert ws['A1'].font.size == 12
    assert ws['A1'].font.bold == True
    assert ws['A1'].fill.start_color.rgb == 'FF993300'

    # Row 2: summary bar
    assert ws['A2'].value == '测试设计'
    assert ws['G2'].value == '测试结果'

    # Row 3: info bar
    assert '模块名称' in str(ws['A3'].value)

    # Row 4: header row
    assert ws['A4'].value == '序号'
    assert ws['B4'].value == '功能模块'
    assert ws['H4'].value == '预期结果'
    assert ws['A4'].fill.start_color.rgb == 'FF9BC2E6'

    # Row 5: section separator
    assert '基金日常管理' in str(ws['A5'].value)
    assert ws['A5'].fill.start_color.rgb == 'FFC0C0C0'

    # Row 6+: data
    assert ws['A6'].value == 'JJRCGL-GSBG_001'
    assert ws['B6'].value == '资源池'
    assert ws['A6'].font.bold == False

    # Column widths
    assert ws.column_dimensions['A'].width >= 18
    assert ws.column_dimensions['H'].width >= 55

    # Borders on data cells
    assert ws['A6'].border.left.style == 'thin'
    assert ws['A6'].border.right.style == 'thin'

    # Merge cells: module (B), sub_module (C), test_point (D) should be merged where same value
    merged = [str(r) for r in ws.merged_cells.ranges]
    # B6:B12 should be merged (all 7 test cases have module="资源池")
    assert any('B6' in m and 'B12' in m for m in merged), f"Expected B6:B12 merged, got: {merged}"
    # C6:C10 should be merged (all "列表页")
    assert any('C6' in m and 'C10' in m for m in merged), f"Expected C6:C10 merged, got: {merged}"

    # COUNTIF formulas in J2, K2, L2
    j2_val = ws['J2'].value
    assert j2_val is not None and 'COUNTIF' in str(j2_val), f"Expected COUNTIF formula in J2, got: {j2_val}"
    k2_val = ws['K2'].value
    assert k2_val is not None and 'COUNTIF' in str(k2_val), f"Expected COUNTIF formula in K2, got: {k2_val}"

    # Row heights for data rows with multi-line content should be > default 15.6
    row6_height = ws.row_dimensions[6].height
    assert row6_height > 15.6, f"Expected row 6 height > 15.6, got: {row6_height}"

    print("All tests passed!")
    os.remove(output)

if __name__ == '__main__':
    test_convert()
