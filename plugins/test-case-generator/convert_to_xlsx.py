#!/usr/bin/env python3
"""
将AI生成的纯文本测试用例转换为格式化的 .xlsx 文件。

用法：
    python convert_to_xlsx.py input.txt output.xlsx [sheet_name]

输入格式：
    每行一条用例，用 | 分隔8列：
    序号|功能模块|子模块|前提条件|测试点|用例名称|操作步骤|预期结果

    分组标题行仅有第1列有值，如：
    列表（ZYC-XZ-L）

    模块标题行仅有第1列有值，如：
    模块公用
"""

import sys
import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 样式定义（匹配现有模板）
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FF000000')
HEADER_FILL = PatternFill(patternType='solid', fgColor='FF95B3D7')
HEADER_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

DATA_FONT = Font(name='微软雅黑', size=10, color='000000')
DATA_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# 所有列定义（与现有模板一致，17列）
HEADERS = [
    '序号', '功能模块', '子模块', '前提条件', '测试点',
    '用例名称', '操作步骤', '预期结果', '',
    '自测结果', '开发人员', '自测日期', '优先级',
    '执行结果', '测试人员', '测试日期', '备注'
]

COL_WIDTHS = {
    'A': 25.5, 'B': 19.6, 'C': 12.1, 'D': 13.4, 'E': 21.9,
    'F': 19.8, 'G': 26.3, 'H': 45.0, 'I': 10.1,
    'J': 10, 'K': 10, 'L': 7.1, 'M': 8, 'N': 10,
    'O': 10, 'P': 10, 'Q': 9
}


def is_group_title(line: str) -> bool:
    """判断是否为分组标题行（如 '列表（ZYC-XZ-L）' 或 '模块公用'）"""
    stripped = line.strip()
    if not stripped:
        return False
    if '|' not in stripped:
        return True
    parts = stripped.split('|')
    return all(p.strip() == '' for p in parts[1:])


def parse_line(line: str) -> list:
    """解析一行数据"""
    stripped = line.strip()
    if not stripped:
        return None

    if is_group_title(stripped):
        title = stripped.split('|')[0].strip()
        return [title] + [None] * 7

    parts = stripped.split('|')
    while len(parts) < 8:
        parts.append('')

    result = []
    for p in parts[:8]:
        v = p.strip()
        result.append(v if v else None)

    return result


def process_newlines(value):
    """将 \\n 转换为真实换行符"""
    if isinstance(value, str):
        return value.replace('\\n', '\n')
    return value


def convert(input_path: str, output_path: str, sheet_name: str = '测试用例'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 设置列宽
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 读取输入文件
    input_file = Path(input_path)
    if not input_file.exists():
        print(f"错误：文件不存在 {input_path}")
        sys.exit(1)

    lines = input_file.read_text(encoding='utf-8').splitlines()

    row_num = 2
    merge_tracking = {
        'B': {'start': None, 'value': None},  # 功能模块
        'C': {'start': None, 'value': None},  # 子模块
        'D': {'start': None, 'value': None},  # 前提条件
        'E': {'start': None, 'value': None},  # 测试点
        'F': {'start': None, 'value': None},  # 用例名称
    }

    def flush_merges():
        """将累计的合并单元格信息写入"""
        for col_letter, info in merge_tracking.items():
            if info['start'] and info['start'] < row_num - 1:
                ws.merge_cells(f"{col_letter}{info['start']}:{col_letter}{row_num - 1}")
            info['start'] = None
            info['value'] = None

    for line in lines:
        parsed = parse_line(line)
        if parsed is None:
            continue

        # 判断是否为分组标题行
        if all(v is None for v in parsed[1:]):
            flush_merges()
            cell = ws.cell(row=row_num, column=1, value=parsed[0])
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            row_num += 1
            continue

        # 写入数据行
        for col_idx in range(8):
            value = process_newlines(parsed[col_idx])
            cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

        # 处理合并逻辑
        col_map = {1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F'}
        for parse_idx, col_letter in col_map.items():
            info = merge_tracking[col_letter]
            if parsed[parse_idx] is not None:
                # 新值出现，先结束之前的合并
                if info['start'] and info['start'] < row_num - 1:
                    ws.merge_cells(f"{col_letter}{info['start']}:{col_letter}{row_num - 1}")
                info['start'] = row_num
                info['value'] = parsed[parse_idx]
            else:
                # 空值，继续合并
                if info['start'] is None:
                    info['start'] = row_num

        row_num += 1

    # 处理最后一批合并
    flush_merges()

    wb.save(output_path)
    print(f"转换完成：{output_path}")
    print(f"共生成 {row_num - 2} 行数据（含分组标题行）")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("用法：python convert_to_xlsx.py input.txt output.xlsx [sheet_name]")
        print()
        print("参数：")
        print("  input.txt    AI生成的纯文本测试用例文件")
        print("  output.xlsx  输出的Excel文件路径")
        print("  sheet_name   工作表名称（可选，默认：测试用例）")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]
    sheet = sys.argv[3] if len(sys.argv) > 3 else '测试用例'

    convert(input_file, output_file, sheet)
